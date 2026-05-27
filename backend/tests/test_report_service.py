from openpyxl import load_workbook

from backend.database import get_db, init_db
from backend.domain.enums import AssetClass, EventType
from backend.services import asset_service, event_service, portfolio_service, report_service


def _b3_import(conn, portfolio_id, filename="2025-12.xlsx"):
    cur = conn.execute(
        """
        INSERT INTO b3_monthly_imports (
            portfolio_id, filename, reference_month, reference_date
        )
        VALUES (?, ?, '2025-12', '2025-12-31')
        """,
        (portfolio_id, filename),
    )
    return cur.lastrowid


def _b3_income(
    conn,
    import_id,
    portfolio_id,
    asset_id,
    *,
    payment_date,
    event_type,
    net_value,
    ledger_event_id=None,
    status="imported",
    source_row=1,
):
    conn.execute(
        """
        INSERT INTO b3_income_events (
            import_id, portfolio_id, asset_id, source_row, payment_date,
            event_type, product, ticker, quantity, unit_price, net_value,
            status, ledger_event_id
        )
        VALUES (?, ?, ?, ?, ?, ?, 'Produto Teste', 'XPTO3', '1', ?, ?, ?, ?)
        """,
        (
            import_id,
            portfolio_id,
            asset_id,
            source_row,
            payment_date,
            event_type,
            net_value,
            net_value,
            status,
            ledger_event_id,
        ),
    )


def test_assets_and_rights_replays_year_end_positions(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        asset = asset_service.create_asset(
            conn,
            AssetClass.ACAO.value,
            "XPTO3",
            market="BR",
            name="XPTO SA",
            cnpj="00.000.000/0001-00",
        )

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2024-01-01", "10", "1000")
        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2025-01-01", "5", "750")
        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.VENDA.value, "2025-06-01", "3", "420")

        report = report_service.list_assets_and_rights(conn, portfolio["id"], 2025)

    assert report["previous_cutoff"] == "2024-12-31"
    assert report["current_cutoff"] == "2025-12-31"
    assert len(report["rows"]) == 1
    row = report["rows"][0]
    assert row["ticker"] == "XPTO3"
    assert row["quantity"] == "12.00"
    assert row["name"] == "XPTO SA"
    assert row["cnpj"] == "00.000.000/0001-00"
    assert row["previous_year_cost"] == "1000.00"
    assert row["current_year_cost"] == "1400.00"


def test_assets_and_rights_includes_fixed_income_classes(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        asset = asset_service.create_asset(conn, AssetClass.TESOURO_DIRETO.value, "TD2029", market="BR")

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2025-01-01", "1", "1000")

        report = report_service.list_assets_and_rights(conn, portfolio["id"], 2025)

    assert len(report["rows"]) == 1
    assert report["rows"][0]["asset_class"] == AssetClass.TESOURO_DIRETO.value
    assert report["rows"][0]["ticker"] == "TD2029"
    assert report["rows"][0]["current_year_cost"] == "1000.00"


def test_income_report_aggregates_eligible_b3_and_ledger_rows(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        acao = asset_service.create_asset(conn, AssetClass.ACAO.value, "XPTO3", market="BR", name="XPTO SA", cnpj="00.000.000/0001-00")
        fii = asset_service.create_asset(conn, AssetClass.FII.value, "FUND11", market="BR", name="FUND FII", cnpj="11.111.111/0001-11")
        etf = asset_service.create_asset(conn, AssetClass.ETF.value, "ETF11", market="BR", name="ETF Fora")
        import_id = _b3_import(conn, portfolio["id"])

        _b3_income(conn, import_id, portfolio["id"], acao["id"], payment_date="2025-02-10", event_type="Dividendo", net_value="100.10", source_row=1)
        _b3_income(conn, import_id, portfolio["id"], acao["id"], payment_date="2025-03-10", event_type="Dividendos", net_value="50.40", source_row=2)
        _b3_income(conn, import_id, portfolio["id"], fii["id"], payment_date="2025-04-10", event_type="Rendimento", net_value="25.00", source_row=3)
        _b3_income(conn, import_id, portfolio["id"], acao["id"], payment_date="2025-05-10", event_type="Juros sobre Capital Próprio", net_value="70.00", source_row=4)
        _b3_income(conn, import_id, portfolio["id"], acao["id"], payment_date="2025-06-10", event_type="Reembolso", net_value="7.50", source_row=5)
        _b3_income(conn, import_id, portfolio["id"], fii["id"], payment_date="2025-06-11", event_type="Reembolso", net_value="2.50", source_row=6)
        _b3_income(conn, import_id, portfolio["id"], etf["id"], payment_date="2025-02-10", event_type="Dividendo", net_value="999.00", source_row=7)
        event_service.create_event(conn, portfolio["id"], acao["id"], EventType.COMPRA.value, "2025-01-01", "10", "100")
        event_service.create_event(conn, portfolio["id"], acao["id"], EventType.BONIFICACAO.value, "2025-08-01", "2", "12.34")

        report = report_service.list_income_report(conn, portfolio["id"], 2025)

    tables = {table["key"]: table for table in report["tables"]}
    exempt_rows = {row["income_type"] + ":" + (row["ticker"] or ""): row for row in tables["tax_exempt"]["rows"]}
    exclusive_rows = {row["income_type"] + ":" + (row["ticker"] or ""): row for row in tables["exclusive_taxation"]["rows"]}

    assert exempt_rows["Dividendo:XPTO3"]["value"] == "150.50"
    assert exempt_rows["Rendimento:FUND11"]["value"] == "25.00"
    assert exempt_rows[f"{EventType.BONIFICACAO.value}:XPTO3"]["value"] == "12.34"
    assert exempt_rows["Reembolso:"]["payer_name"] == report_service.REIMBURSEMENT_PAYER_NAME
    assert exempt_rows["Reembolso:"]["value"] == "10.00"
    assert exclusive_rows["Juros Sobre Capital Próprio:XPTO3"]["value"] == "70.00"
    assert tables["tax_exempt"]["total"] == "197.84"
    assert tables["exclusive_taxation"]["total"] == "70.00"


def test_income_report_adds_exempt_capital_gain_lines_without_asset_binding(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        stock = asset_service.create_asset(conn, AssetClass.ACAO.value, "ACAO3", market="BR")
        fi_infra = asset_service.create_asset(
            conn,
            AssetClass.FI_INFRA.value,
            "INFRA11",
            market="BR",
        )

        event_service.create_event(conn, portfolio["id"], stock["id"], EventType.COMPRA.value, "2025-01-02", "100", "10000")
        event_service.create_event(conn, portfolio["id"], stock["id"], EventType.VENDA.value, "2025-02-20", "100", "11000", gross_value="11000")
        event_service.create_event(conn, portfolio["id"], fi_infra["id"], EventType.COMPRA.value, "2025-03-01", "100", "10000")
        event_service.create_event(conn, portfolio["id"], fi_infra["id"], EventType.VENDA.value, "2025-04-20", "100", "12000", gross_value="12000")

        report = report_service.list_income_report(conn, portfolio["id"], 2025)

    tax_exempt = next(table for table in report["tables"] if table["key"] == "tax_exempt")
    rows_by_name = {row["payer_name"]: row for row in tax_exempt["rows"]}

    stock_row = rows_by_name[report_service.STOCK_EXEMPT_GAIN_PAYER_NAME]
    fi_infra_row = rows_by_name[report_service.FI_INFRA_EXEMPT_GAIN_PAYER_NAME]

    assert stock_row["ticker"] is None
    assert stock_row["payer_cnpj"] is None
    assert stock_row["income_type"] == report_service.CAPITAL_GAIN_EXEMPT_INCOME_TYPE
    assert stock_row["value"] == "1000.00"
    assert fi_infra_row["ticker"] is None
    assert fi_infra_row["payer_cnpj"] is None
    assert fi_infra_row["income_type"] == report_service.CAPITAL_GAIN_EXEMPT_INCOME_TYPE
    assert fi_infra_row["value"] == "2000.00"
    assert tax_exempt["total"] == "3000.00"


def test_income_report_resolves_cancelled_and_duplicate_amortizations(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        asset = asset_service.create_asset(conn, AssetClass.FII.value, "FUND11", market="BR", name="FUND FII", cnpj="11.111.111/0001-11")
        import_id = _b3_import(conn, portfolio["id"])

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2024-01-01", "100", "1000")
        manual = event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-05-30", "0", "271.81")
        duplicate = event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-05-30", "0", "271.81")
        conn.execute("UPDATE events SET duplicate_flag = 1 WHERE id = ?", (duplicate["id"],))
        _b3_income(
            conn,
            import_id,
            portfolio["id"],
            asset["id"],
            payment_date="2025-05-30",
            event_type="Amortização",
            net_value="271.81",
            ledger_event_id=duplicate["id"],
            status="ledger_event_created",
        )
        _b3_income(
            conn,
            import_id,
            portfolio["id"],
            asset["id"],
            payment_date="2025-05-30",
            event_type="Amortização",
            net_value="271.81",
            source_row=2,
        )

        report = report_service.list_income_report(conn, portfolio["id"], 2025)
        conn.execute("UPDATE events SET is_cancelled = 1 WHERE id = ?", (manual["id"],))
        conn.execute("UPDATE events SET duplicate_flag = 0 WHERE id = ?", (duplicate["id"],))
        confirmed_report = report_service.list_income_report(conn, portfolio["id"], 2025)

    rows = {row["income_type"]: row for row in report["tables"][0]["rows"]}
    confirmed_rows = {row["income_type"]: row for row in confirmed_report["tables"][0]["rows"]}

    assert rows[EventType.AMORTIZACAO.value]["value"] == "271.81"
    assert confirmed_rows[EventType.AMORTIZACAO.value]["value"] == "271.81"


def test_income_report_keeps_active_side_of_duplicate_amortizations(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        asset = asset_service.create_asset(conn, AssetClass.FII.value, "FUND11", market="BR", name="FUND FII", cnpj="11.111.111/0001-11")
        import_id = _b3_import(conn, portfolio["id"])

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2024-01-01", "100", "1000")

        manual_cancelled = event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-06-30", "0", "100.00")
        duplicate_only_active = event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-06-30", "0", "100.00")
        conn.execute("UPDATE events SET is_cancelled = 1 WHERE id = ?", (manual_cancelled["id"],))
        conn.execute("UPDATE events SET duplicate_flag = 1 WHERE id = ?", (duplicate_only_active["id"],))
        _b3_income(
            conn,
            import_id,
            portfolio["id"],
            asset["id"],
            payment_date="2025-06-30",
            event_type=EventType.AMORTIZACAO.value,
            net_value="100.00",
            ledger_event_id=duplicate_only_active["id"],
            status="ledger_event_created",
            source_row=1,
        )

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-07-31", "0", "200.00")
        duplicate_cancelled = event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-07-31", "0", "200.00")
        conn.execute("UPDATE events SET duplicate_flag = 1, is_cancelled = 1 WHERE id = ?", (duplicate_cancelled["id"],))
        _b3_income(
            conn,
            import_id,
            portfolio["id"],
            asset["id"],
            payment_date="2025-07-31",
            event_type=EventType.AMORTIZACAO.value,
            net_value="200.00",
            ledger_event_id=duplicate_cancelled["id"],
            status="ledger_event_created",
            source_row=2,
        )

        both_manual_cancelled = event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-08-29", "0", "300.00")
        both_duplicate_cancelled = event_service.create_event(conn, portfolio["id"], asset["id"], EventType.AMORTIZACAO.value, "2025-08-29", "0", "300.00")
        conn.execute("UPDATE events SET is_cancelled = 1 WHERE id = ?", (both_manual_cancelled["id"],))
        conn.execute("UPDATE events SET duplicate_flag = 1, is_cancelled = 1 WHERE id = ?", (both_duplicate_cancelled["id"],))
        _b3_income(
            conn,
            import_id,
            portfolio["id"],
            asset["id"],
            payment_date="2025-08-29",
            event_type=EventType.AMORTIZACAO.value,
            net_value="300.00",
            ledger_event_id=both_duplicate_cancelled["id"],
            status="ledger_event_created",
            source_row=3,
        )

        _b3_income(
            conn,
            import_id,
            portfolio["id"],
            asset["id"],
            payment_date="2025-09-30",
            event_type="Rendimento",
            net_value="25.00",
            source_row=4,
        )

        report = report_service.list_income_report(conn, portfolio["id"], 2025)

    rows = {row["income_type"]: row for row in report["tables"][0]["rows"]}

    assert rows[EventType.AMORTIZACAO.value]["value"] == "300.00"
    assert rows["Rendimento"]["value"] == "25.00"


def test_fiscal_report_xlsx_uses_assets_and_rights_section(tmp_path):
    db_path = tmp_path / "ledger.db"
    xlsx_path = tmp_path / "report.xlsx"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        asset = asset_service.create_asset(conn, AssetClass.FII.value, "ABCD11", market="BR", name="ABCD FII")

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2025-01-01", "10", "950")

        xlsx_path.write_bytes(
            report_service.build_fiscal_report_xlsx(
                conn,
                portfolio["id"],
                2025,
                sections=["assets_and_rights"],
            )
        )

    workbook = load_workbook(xlsx_path)
    worksheet = workbook["Bens e Direitos"]

    assert worksheet["A1"].value == "CLASSE"
    assert worksheet["B1"].value == "TICKER"
    assert worksheet["D1"].value == "NOME ATIVO"
    assert worksheet["F1"].value == "Situação em 31/12/2024"
    assert worksheet["G1"].value == "Situação em 31/12/2025"
    assert worksheet["A2"].value == AssetClass.FII.value
    assert worksheet["B2"].value == "ABCD11"
    assert worksheet["C2"].value == 10
    assert worksheet["D2"].value == "ABCD FII"
    assert worksheet["F2"].value == 0
    assert worksheet["G2"].value == 950
    assert worksheet["C2"].number_format == "#,##0.00########"
    assert worksheet["F2"].number_format == "#,##0.00"
    assert worksheet["G2"].number_format == "#,##0.00"


def test_assets_and_rights_xlsx_delegates_to_fiscal_export(tmp_path):
    db_path = tmp_path / "ledger.db"
    xlsx_path = tmp_path / "report.xlsx"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        asset = asset_service.create_asset(conn, AssetClass.FII.value, "ABCD11", market="BR", name="ABCD FII")

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2025-01-01", "10", "950")

        xlsx_path.write_bytes(report_service.build_assets_and_rights_xlsx(conn, portfolio["id"], 2025))

    workbook = load_workbook(xlsx_path)

    assert workbook.sheetnames == ["Bens e Direitos"]
    assert workbook["Bens e Direitos"]["G2"].value == 950


def test_fiscal_report_xlsx_defaults_to_all_annual_sections(tmp_path):
    db_path = tmp_path / "ledger.db"
    xlsx_path = tmp_path / "report.xlsx"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal")
        asset = asset_service.create_asset(conn, AssetClass.ACAO.value, "XPTO3", market="BR", name="XPTO SA")
        import_id = _b3_import(conn, portfolio["id"])

        event_service.create_event(conn, portfolio["id"], asset["id"], EventType.COMPRA.value, "2025-01-01", "10", "100")
        _b3_income(conn, import_id, portfolio["id"], asset["id"], payment_date="2025-03-10", event_type="Dividendo", net_value="123.45")

        xlsx_path.write_bytes(report_service.build_fiscal_report_xlsx(conn, portfolio["id"], 2025))

    workbook = load_workbook(xlsx_path)

    assert workbook.sheetnames == ["Bens e Direitos", "Rendimentos Isentos", "Tributação Exclusiva"]
    assert workbook["Rendimentos Isentos"]["A1"].value == "TICKER"
    assert workbook["Rendimentos Isentos"]["B1"].value == "CNPJ DA FONTE PAGADORA"
    assert workbook["Rendimentos Isentos"]["C1"].value == "NOME DA FONTE PAGADORA"
    assert workbook["Rendimentos Isentos"]["D1"].value == "TIPO"
    assert workbook["Rendimentos Isentos"]["E1"].value == "VALOR"
    assert workbook["Rendimentos Isentos"]["D2"].value == "Dividendo"
    assert workbook["Rendimentos Isentos"]["E2"].value == 123.45
    assert workbook["Rendimentos Isentos"]["D3"].value is None
    assert workbook["Rendimentos Isentos"]["E3"].value is None
    assert workbook["Rendimentos Isentos"]["E2"].number_format == "#,##0.00"
    assert workbook[workbook.sheetnames[2]]["D2"].value is None
    assert workbook[workbook.sheetnames[2]]["E2"].value is None


def test_fiscal_report_filename_slugifies_portfolio_name():
    assert (
        report_service.fiscal_report_filename(2026, "Nome Carteira")
        == "relatorio-fiscal-2026-nome_carteira.xlsx"
    )
    assert (
        report_service.fiscal_report_filename(2026, "Carteira Ágil / João")
        == "relatorio-fiscal-2026-carteira_agil_joao.xlsx"
    )
