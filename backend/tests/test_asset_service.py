import json
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from backend.database import get_db, init_db
from backend.domain.enums import AssetClass, TreasuryIndexer
from backend.main import app
from backend.services import asset_service, portfolio_service
from backend.services.import_service import import_to_ledger
from backend.services.portfolio_service import create_portfolio


def test_delete_merged_target_clears_merge_reference(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        source = asset_service.create_asset(conn, AssetClass.ACAO.value, "AAAA3", market="BR")
        target = asset_service.create_asset(conn, AssetClass.ACAO.value, "BBBB3", market="BR")

        asset_service.merge_assets(conn, source["id"], target["id"])
        assert asset_service.delete_asset(conn, target["id"])

        assert asset_service.get_asset(conn, target["id"]) is None
        unmerged_source = asset_service.get_asset(conn, source["id"])
        assert unmerged_source["merged_into_asset_id"] is None
        assert unmerged_source["merged_at"] is None


def test_asset_service_persists_fiscal_fields_on_create_update_and_response(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        asset = asset_service.create_asset(
            conn,
            AssetClass.FI_INFRA.value,
            "INFRA11",
            market="BR",
            fiscal_regime_override="FI_INFRA_EXEMPT",
            fiscal_tax_treatment="EXEMPT_ZERO",
        )
        assert asset["fiscal_regime_override"] == "FI_INFRA_EXEMPT"
        assert asset["fiscal_tax_treatment"] == "EXEMPT_ZERO"

        updated = asset_service.update_asset_metadata(
            conn,
            asset["id"],
            fiscal_regime_override="B3_COMMON_15",
            fiscal_tax_treatment="TAXABLE",
        )

        assert updated["fiscal_regime_override"] == "B3_COMMON_15"
        assert updated["fiscal_tax_treatment"] == "TAXABLE"
        assert asset_service.get_asset(conn, asset["id"])["fiscal_tax_treatment"] == "TAXABLE"


def test_asset_service_normalizes_cnpj_on_create_update_and_response(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        asset = asset_service.create_asset(
            conn,
            AssetClass.ACAO.value,
            "XPTO3",
            market="BR",
            cnpj="00.000.000/0001-00",
        )
        assert asset["cnpj"] == "00000000000100"

        updated = asset_service.update_asset_metadata(
            conn,
            asset["id"],
            cnpj="11.111.111/0001-11",
        )

        assert updated["cnpj"] == "11111111000111"
        assert asset_service.get_asset(conn, asset["id"])["cnpj"] == "11111111000111"

        with pytest.raises(ValueError, match="CNPJ"):
            asset_service.update_asset_metadata(conn, asset["id"], cnpj="123")


@pytest.mark.parametrize("asset_class", ["AÃ§Ã£o", "A��o", "Acao", "acao", "AÇÃO", "Debenture", "Deb�nture"])
def test_manual_asset_class_rejects_aliases_and_mojibake_without_persisting(tmp_path, asset_class):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        with pytest.raises(ValueError, match="Classe do ativo invalido"):
            asset_service.create_asset(conn, asset_class, "XPTO3", market="BR")

        count = conn.execute("SELECT COUNT(*) AS count FROM assets").fetchone()["count"]

    assert count == 0


def test_xlsx_import_maps_safe_asset_class_aliases_to_canonical_values(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registro"
    sheet.append(["Classe", "Ativo", "Evento", "Data", "Quantidade", "Valor Evento", "Valor Bruto"])
    sheet.append(["Acao", " xpto3 ", "Compra", "2026-05-10", 10, 1000, None])
    sheet.append(["Debenture", "deb123", "Compra", "2026-05-11", 1, 100, None])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    with get_db(db_path) as conn:
        portfolio = create_portfolio(conn, "Principal")
        result = import_to_ledger(conn, stream, portfolio["id"])
        rows = conn.execute(
            """
            SELECT a.asset_class, t.ticker
            FROM assets a
            JOIN asset_tickers t ON t.asset_id = a.id
            ORDER BY a.id
            """
        ).fetchall()

    assert result["imported"] == 2
    assert [(row["asset_class"], row["ticker"]) for row in rows] == [
        (AssetClass.ACAO.value, "XPTO3"),
        (AssetClass.DEBENTURE.value, "DEB123"),
    ]


def test_ticker_is_uppercase_and_exact_ticker_queries_use_normalized_parameter(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        asset = asset_service.create_asset(conn, AssetClass.ACAO.value, " xpto3 ", market="BR")
        ticker_row = conn.execute("SELECT ticker FROM asset_tickers WHERE asset_id = ?", (asset["id"],)).fetchone()
        resolved_id = asset_service.resolve_ticker_to_asset_id(conn, "xpto3", "2026-05-10")
        found_id = asset_service.find_asset_by_ticker(conn, "xpto3")
        match = asset_service.match_asset(conn, "xpto3", AssetClass.ACAO.value, "2026-05-10", "BR")

    assert ticker_row["ticker"] == "XPTO3"
    assert resolved_id == asset["id"]
    assert found_id == asset["id"]
    assert match["status"] == "exact"
    assert match["asset"]["id"] == asset["id"]


def test_asset_service_rejects_invalid_market_currency_and_incompatible_metadata(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        with pytest.raises(ValueError, match="Mercado"):
            asset_service.create_asset(conn, AssetClass.ACAO.value, "BADM3", market="EU")
        with pytest.raises(ValueError, match="Moeda"):
            asset_service.create_asset(conn, AssetClass.ACAO.value, "BADC3", currency="EUR")
        with pytest.raises(ValueError, match="Metadados incompativeis"):
            asset_service.create_asset(conn, AssetClass.FII.value, "META11", market="BR", gics_sector="Real Estate")

        asset = asset_service.create_asset(conn, AssetClass.ACAO.value, "XPTO3", market="BR")
        with pytest.raises(ValueError, match="Metadados incompativeis"):
            asset_service.update_asset_metadata(conn, asset["id"], gics_sector="Technology")

        persisted = conn.execute("SELECT COUNT(*) AS count FROM assets").fetchone()["count"]

    assert persisted == 1


def test_review_status_and_portfolio_boolean_are_normalized_or_rejected(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        portfolio = portfolio_service.create_portfolio(conn, "Principal", consolidated="1")
        review = asset_service.create_match_review(conn, "test", "xpto3", AssetClass.ACAO.value, "BR")
        resolved = asset_service.resolve_match_review(conn, review["id"])

        assert portfolio["consolidated"] == 1
        assert review["status"] == "pending"
        assert resolved["status"] == "resolved"

        with pytest.raises(ValueError, match="Status de revisao"):
            asset_service.list_match_reviews(conn, "PENDING")
        with pytest.raises(ValueError, match="Booleano"):
            portfolio_service.create_portfolio(conn, "Invalida", consolidated=2)


def test_asset_api_contract_exposes_fiscal_fields_on_create_get_and_patch(tmp_path, monkeypatch):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)
    monkeypatch.setattr("backend.routers.assets.get_db", lambda: get_db(db_path))
    client = TestClient(app)

    create_response = client.post(
        "/api/assets",
        json={
            "asset_class": AssetClass.FI_INFRA.value,
            "ticker": "INFRA11",
            "market": "BR",
            "fiscal_regime_override": "FI_INFRA_EXEMPT",
            "fiscal_tax_treatment": "EXEMPT_ZERO",
        },
    )

    assert create_response.status_code == 200
    created = create_response.json()
    assert created["fiscal_regime_override"] == "FI_INFRA_EXEMPT"
    assert created["fiscal_tax_treatment"] == "EXEMPT_ZERO"

    patch_response = client.patch(
        f"/api/assets/{created['id']}",
        json={
            "fiscal_regime_override": "B3_COMMON_15",
            "fiscal_tax_treatment": "TAXABLE",
        },
    )
    assert patch_response.status_code == 200
    patched = patch_response.json()
    assert patched["fiscal_regime_override"] == "B3_COMMON_15"
    assert patched["fiscal_tax_treatment"] == "TAXABLE"

    get_response = client.get(f"/api/assets/{created['id']}")
    assert get_response.status_code == 200
    fetched = get_response.json()
    assert fetched["fiscal_regime_override"] == "B3_COMMON_15"
    assert fetched["fiscal_tax_treatment"] == "TAXABLE"

    unsupported_response = client.post(
        "/api/assets",
        json={
            "asset_class": AssetClass.ACAO.value,
            "ticker": "FOREIGN3",
            "market": "BR",
            "fiscal_regime_override": "FOREIGN_14754",
        },
    )
    assert unsupported_response.status_code == 400


def test_asset_service_persists_and_validates_treasury_indexer(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        treasury = asset_service.create_asset(
            conn,
            AssetClass.TESOURO_DIRETO.value,
            "TESOURO-SELIC-2029",
            treasury_indexer=TreasuryIndexer.SELIC.value,
        )

        assert treasury["treasury_indexer"] == "SELIC"

        updated = asset_service.update_asset_metadata(
            conn,
            treasury["id"],
            treasury_indexer=TreasuryIndexer.PREFIXED.value,
        )
        assert updated["treasury_indexer"] == "PREFIXED"

        cleared = asset_service.update_asset_metadata(conn, treasury["id"], treasury_indexer="")
        assert cleared["treasury_indexer"] is None

        with pytest.raises(ValueError, match="Indexador"):
            asset_service.create_asset(
                conn,
                AssetClass.TESOURO_DIRETO.value,
                "TESOURO-LIVRE-2030",
                treasury_indexer="CDI",
            )

        with pytest.raises(ValueError, match="Tesouro Direto"):
            asset_service.create_asset(
                conn,
                AssetClass.ACAO.value,
                "INDEX3",
                treasury_indexer=TreasuryIndexer.IPCA.value,
            )


def test_asset_api_contract_exposes_treasury_indexer(tmp_path, monkeypatch):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)
    monkeypatch.setattr("backend.routers.assets.get_db", lambda: get_db(db_path))
    client = TestClient(app)

    create_response = client.post(
        "/api/assets",
        json={
            "asset_class": AssetClass.TESOURO_DIRETO.value,
            "ticker": "TESOURO-IPCA-2035",
            "treasury_indexer": "IPCA",
        },
    )

    assert create_response.status_code == 200
    created = create_response.json()
    assert created["treasury_indexer"] == "IPCA"

    patch_response = client.patch(
        f"/api/assets/{created['id']}",
        json={"treasury_indexer": "PREFIXED"},
    )
    assert patch_response.status_code == 200
    assert patch_response.json()["treasury_indexer"] == "PREFIXED"

    get_response = client.get(f"/api/assets/{created['id']}")
    assert get_response.status_code == 200
    assert get_response.json()["treasury_indexer"] == "PREFIXED"

    invalid_response = client.patch(f"/api/assets/{created['id']}", json={"treasury_indexer": "CDI"})
    assert invalid_response.status_code == 400


def test_asset_service_persists_international_taxonomy_fields(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        stock = asset_service.create_asset(
            conn,
            AssetClass.STOCK.value,
            "AAPL",
            gics_sector="Information Technology",
            gics_industry_group="Technology Hardware & Equipment",
            gics_industry="Technology Hardware, Storage & Peripherals",
            gics_sub_industry="Technology Hardware, Storage & Peripherals",
        )

        assert stock["sector"] is None
        assert stock["subsector"] is None
        assert stock["segment"] is None
        assert stock["gics_sector"] == "Information Technology"
        assert stock["gics_industry_group"] == "Technology Hardware & Equipment"
        assert stock["gics_industry"] == "Technology Hardware, Storage & Peripherals"
        assert stock["gics_sub_industry"] == "Technology Hardware, Storage & Peripherals"

        updated_stock = asset_service.update_asset_metadata(
            conn,
            stock["id"],
            gics_sector="Communication Services",
            gics_industry_group="Media & Entertainment",
        )

        assert updated_stock["gics_sector"] == "Communication Services"
        assert updated_stock["gics_industry_group"] == "Media & Entertainment"

        reit = asset_service.create_asset(
            conn,
            AssetClass.REIT.value,
            "PLD",
            reit_type="Equity",
            gics_sector="Real Estate",
            gics_industry_group="Equity Real Estate Investment Trusts (REITs)",
        )

        assert reit["reit_type"] == "Equity"
        assert reit["gics_sector"] == "Real Estate"

        updated_reit = asset_service.update_asset_metadata(conn, reit["id"], reit_type="Mortgage")
        assert updated_reit["reit_type"] == "Mortgage"

        cleared_reit = asset_service.update_asset_metadata(conn, reit["id"], reit_type="")
        assert cleared_reit["reit_type"] is None

        with pytest.raises(ValueError, match="REIT Type"):
            asset_service.update_asset_metadata(conn, reit["id"], reit_type="Office")


def test_asset_api_contract_exposes_international_taxonomy_fields(tmp_path, monkeypatch):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)
    monkeypatch.setattr("backend.routers.assets.get_db", lambda: get_db(db_path))
    client = TestClient(app)

    create_response = client.post(
        "/api/assets",
        json={
            "asset_class": AssetClass.REIT.value,
            "ticker": "PLD",
            "reit_type": "Equity",
            "gics_sector": "Real Estate",
            "gics_industry_group": "Equity Real Estate Investment Trusts (REITs)",
            "gics_industry": "Industrial REITs",
            "gics_sub_industry": "Industrial REITs",
        },
    )

    assert create_response.status_code == 200
    created = create_response.json()
    assert created["reit_type"] == "Equity"
    assert created["gics_sector"] == "Real Estate"
    assert created["gics_industry_group"] == "Equity Real Estate Investment Trusts (REITs)"
    assert created["gics_industry"] == "Industrial REITs"
    assert created["gics_sub_industry"] == "Industrial REITs"

    patch_response = client.patch(
        f"/api/assets/{created['id']}",
        json={
            "reit_type": "Hybrid",
            "gics_industry": "Diversified REITs",
            "gics_sub_industry": "Diversified REITs",
        },
    )

    assert patch_response.status_code == 200
    patched = patch_response.json()
    assert patched["reit_type"] == "Hybrid"
    assert patched["gics_industry"] == "Diversified REITs"
    assert patched["gics_sub_industry"] == "Diversified REITs"

    get_response = client.get(f"/api/assets/{created['id']}")
    assert get_response.status_code == 200
    fetched = get_response.json()
    assert fetched["reit_type"] == "Hybrid"
    assert fetched["gics_sector"] == "Real Estate"

    invalid_response = client.patch(f"/api/assets/{created['id']}", json={"reit_type": "Office"})
    assert invalid_response.status_code == 400


def test_probable_match_review_stores_operation_payload(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        asset_service.create_asset(conn, AssetClass.ACAO.value, "XPTO3", market="BR")

        with pytest.raises(ValueError, match="revis"):
            asset_service.create_asset(
                conn,
                "FII",
                "XPTO3",
                market="BR",
                event_date="2026-05-11",
                portfolio_id=7,
                event_type="Compra",
                quantity="10",
                event_value="1000",
                notes="teste",
                source="event_form",
            )

        reviews = asset_service.list_match_reviews(conn)
        assert len(reviews) == 1
        payload = json.loads(reviews[0]["operation_payload"])
        assert payload["ticker"] == "XPTO3"
        assert payload["asset_class"] == "FII"
        assert payload["market"] == "BR"
        assert payload["portfolio_id"] == 7
        assert payload["event_type"] == "Compra"
        assert payload["event_date"] == "2026-05-11"
        assert payload["quantity"] == "10"
        assert payload["event_value"] == "1000"
        assert payload["notes"] == "teste"


def test_probable_match_review_reuses_existing_and_preserves_first_payload(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        asset_service.create_asset(conn, AssetClass.ACAO.value, "XPTO3", market="BR")

        for quantity in ["10", "20"]:
            with pytest.raises(ValueError):
                asset_service.create_asset(
                    conn,
                    "FII",
                    "XPTO3",
                    market="BR",
                    event_date="2026-05-11",
                    portfolio_id=1,
                    event_type="Compra",
                    quantity=quantity,
                    event_value="1000",
                    source="event_form",
                )

        reviews = asset_service.list_match_reviews(conn)
        assert len(reviews) == 1
        payload = json.loads(reviews[0]["operation_payload"])
        assert payload["quantity"] == "10"


def test_list_match_reviews_filters_by_operation_payload_portfolio(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    with get_db(db_path) as conn:
        first = asset_service.create_match_review(
            conn,
            source="event_form",
            ticker="AAAA3",
            asset_class=AssetClass.ACAO.value,
            market="BR",
            event_date="2026-05-11",
            reason="Revisao portfolio 1",
            operation_payload={"portfolio_id": 1},
        )
        second = asset_service.create_match_review(
            conn,
            source="event_form",
            ticker="BBBB3",
            asset_class=AssetClass.ACAO.value,
            market="BR",
            event_date="2026-05-12",
            reason="Revisao portfolio 2",
            operation_payload={"portfolio_id": 2},
        )
        without_portfolio = asset_service.create_match_review(
            conn,
            source="manual",
            ticker="CCCC3",
            asset_class=AssetClass.ACAO.value,
            market="BR",
            event_date="2026-05-13",
            reason="Revisao sem portfolio",
        )

        all_reviews = asset_service.list_match_reviews(conn)
        assert {review["id"] for review in all_reviews} == {first["id"], second["id"], without_portfolio["id"]}

        first_portfolio_reviews = asset_service.list_match_reviews(conn, portfolio_id=1)
        assert [review["id"] for review in first_portfolio_reviews] == [first["id"]]

        second_portfolio_reviews = asset_service.list_match_reviews(conn, portfolio_id=2)
        assert [review["id"] for review in second_portfolio_reviews] == [second["id"]]


def test_import_review_stores_operation_payload(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registro"
    sheet.append(["Classe", "Ativo", "Evento", "Data", "Quantidade", "Valor Evento", "Valor Bruto"])
    sheet.append(["FII", "XPTO3", "Venda", "2026-05-11", 10, 990, 1000])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    with get_db(db_path) as conn:
        portfolio = create_portfolio(conn, "Principal")
        asset_service.create_asset(conn, AssetClass.ACAO.value, "XPTO3", market="BR")

        result = import_to_ledger(conn, stream, portfolio["id"])

        assert result["imported"] == 0
        assert result["skipped"] == 1
        assert result["review_count"] == 1

        reviews = asset_service.list_match_reviews(conn)
        assert len(reviews) == 1
        payload = json.loads(reviews[0]["operation_payload"])
        assert payload["ticker"] == "XPTO3"
        assert payload["asset_class"] == "FII"
        assert payload["portfolio_id"] == portfolio["id"]
        assert payload["event_type"] == "Venda"
        assert payload["event_date"] == "2026-05-11"
        assert payload["quantity"] == "10"
        assert payload["event_value"] == "990"
        assert payload["gross_value"] == "1000"
        assert payload["source_row"] == 2


def test_import_xlsx_persists_gross_value_only_for_sales(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registro"
    sheet.append(["Classe", "Ativo", "Evento", "Data", "Quantidade", "Valor Evento", "Valor Bruto"])
    sheet.append([AssetClass.ACAO.value, "XPTO3", "Compra", "2026-05-10", 10, 1000, 1005])
    sheet.append([AssetClass.ACAO.value, "XPTO3", "Venda", "2026-05-11", 4, 390, 400])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    with get_db(db_path) as conn:
        portfolio = create_portfolio(conn, "Principal")
        result = import_to_ledger(conn, stream, portfolio["id"])
        rows = conn.execute("SELECT * FROM events ORDER BY id").fetchall()

    assert result["imported"] == 2
    assert rows[0]["event_type"] == "Compra"
    assert rows[0]["gross_value"] is None
    assert rows[1]["event_type"] == "Venda"
    assert rows[1]["event_value"] == "390"
    assert rows[1]["gross_value"] == "400"


def test_import_xlsx_duplicate_preserves_sale_gross_value(tmp_path):
    db_path = tmp_path / "ledger.db"
    init_db(db_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registro"
    sheet.append(["Classe", "Ativo", "Evento", "Data", "Quantidade", "Valor Evento", "Valor Bruto"])
    sheet.append([AssetClass.ACAO.value, "XPTO3", "Compra", "2026-05-10", 10, 1000, None])
    sheet.append([AssetClass.ACAO.value, "XPTO3", "Venda", "2026-05-11", 4, 390, 400])
    stream = BytesIO()
    workbook.save(stream)

    with get_db(db_path) as conn:
        portfolio = create_portfolio(conn, "Principal")
        stream.seek(0)
        first = import_to_ledger(conn, stream, portfolio["id"])
        stream.seek(0)
        second = import_to_ledger(conn, stream, portfolio["id"])
        rows = conn.execute("SELECT * FROM events ORDER BY id").fetchall()

    assert first["duplicates"] == 0
    assert second["duplicates"] == 2
    assert rows[3]["event_type"] == "Venda"
    assert rows[3]["duplicate_flag"] == 1
    assert rows[3]["gross_value"] == "400"
