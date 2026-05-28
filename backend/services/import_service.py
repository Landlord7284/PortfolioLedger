"""
XLSX importer for event data.

This module handles importing Excel spreadsheets into the ledger.
Formula support is intentionally limited to **simple arithmetic
expressions** (addition, subtraction, multiplication, division) as a legacy
compatibility feature.  Complex formulas, cell references, and Excel functions
are rejected with explicit errors.

After import, only the final computed values are persisted in the ledger.
"""

from __future__ import annotations

import ast
import operator
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Optional, BinaryIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from backend.domain.enums import AssetClass, EventType
from backend.domain.engine import to_decimal


# ─────────────────────────────────────────────────────────────
# Safe formula evaluator (legacy support only)
# ─────────────────────────────────────────────────────────────

# Allowed operators for the safe evaluator
_SAFE_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
}

# Pattern that detects cell references (e.g., A1, $B$2, Sheet1!A1)
_CELL_REF_RE = re.compile(r"(?<![A-Za-z])(\$?[A-Za-z]{1,3}\$?\d+|[A-Za-z]+!\$?[A-Za-z]+\$?\d+)")
# Pattern that detects Excel function calls (e.g., SUM(...), VLOOKUP(...))
_FUNC_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_.]*\s*\(")


def _safe_eval_formula(formula: str) -> Decimal:
    """
    Evaluate a simple arithmetic Excel formula and return a Decimal.

    Only supports: +, -, *, / with numeric literals.
    Rejects cell references, function calls, and any other complex constructs.

    Raises ValueError for unsupported formulas.
    """
    # Strip the leading '='
    expr = formula.lstrip("=").strip()

    # Reject cell references
    if _CELL_REF_RE.search(expr):
        raise ValueError(
            f"Fórmula contém referência a célula e não pode ser resolvida: '{formula}'"
        )

    # Reject function calls
    if _FUNC_RE.search(expr):
        raise ValueError(
            f"Fórmula contém funções Excel e não pode ser resolvida: '{formula}'"
        )

    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError:
        raise ValueError(f"Fórmula com sintaxe inválida: '{formula}'")

    result = _eval_node(tree.body)
    return Decimal(str(result))


def _eval_node(node: ast.AST) -> float:
    """Recursively evaluate an AST node (numbers and basic ops only)."""
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    elif isinstance(node, ast.BinOp):
        op_func = _SAFE_OPS.get(type(node.op))
        if op_func is None:
            raise ValueError(f"Operador não suportado: {type(node.op).__name__}")
        left = _eval_node(node.left)
        right = _eval_node(node.right)
        return op_func(left, right)
    elif isinstance(node, ast.UnaryOp):
        op_func = _SAFE_OPS.get(type(node.op))
        if op_func is None:
            raise ValueError(f"Operador unário não suportado: {type(node.op).__name__}")
        return op_func(_eval_node(node.operand))
    else:
        raise ValueError(
            f"Expressão não suportada na fórmula: {ast.dump(node)}"
        )


def _resolve_cell_value(value, cell_ref: str = "") -> Decimal:
    """
    Resolve a cell value to Decimal.

    - If the value is a string starting with '=', treat as formula.
    - Otherwise, convert directly.
    """
    if value is None:
        return Decimal("0")

    if isinstance(value, str):
        s = value.strip()
        if s.startswith("="):
            return _safe_eval_formula(s)
        return to_decimal(s)

    return to_decimal(value)


# ─────────────────────────────────────────────────────────────
# Class label mapping (legacy XLSX → enum)
# ─────────────────────────────────────────────────────────────

_CLASS_LABEL_MAP = {
    "Ação": AssetClass.ACAO,
    "Acao": AssetClass.ACAO,
    "BDR": AssetClass.BDR,
    "Criptomoeda": AssetClass.CRIPTOMOEDA,
    "Debênture": AssetClass.DEBENTURE,
    "Debenture": AssetClass.DEBENTURE,
    "CRI": AssetClass.CRI,
    "CRA": AssetClass.CRA,
    "ETF": AssetClass.ETF,
    "FII": AssetClass.FII,
    "FI-INFRA": AssetClass.FI_INFRA,
    "Tesouro Direto": AssetClass.TESOURO_DIRETO,
    "TD": AssetClass.TESOURO_DIRETO,
    "Stock": AssetClass.STOCK,
    "REIT": AssetClass.REIT,
}

_EVENT_LABEL_MAP = {
    "Compra": EventType.COMPRA,
    "Venda": EventType.VENDA,
    "Desdobramento": EventType.DESDOBRAMENTO,
    "Grupamento": EventType.GRUPAMENTO,
    "Amortização": EventType.AMORTIZACAO,
    "Amortizacao": EventType.AMORTIZACAO,
    "Bonificação": EventType.BONIFICACAO,
    "Bonificacao": EventType.BONIFICACAO,
    "Cisão": EventType.CISAO,
    "Cisao": EventType.CISAO,
    "Resgate Antecipado": EventType.RESGATE_ANTECIPADO,
    "Resgate Vencimento": EventType.RESGATE_VENCIMENTO,
}


IMPORT_TEMPLATE_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_BR_TEMPLATE_HEADERS = [
    "Classe",
    "Ativo",
    "Evento",
    "Data",
    "Quantidade",
    "Valor Evento",
    "Valor Bruto",
]

_INTERNATIONAL_TEMPLATE_HEADERS = [
    "Classe",
    "Ativo",
    "Evento",
    "Data",
    "Quantidade",
    "Valor Evento",
    "Origem US",
]

_BR_TEMPLATE_INSTRUCTIONS = [
    ("Classe", "Use valores como Ação, BDR, ETF, FII, FI-INFRA, Tesouro Direto, Debênture."),
    ("Ativo", "Ticker do ativo, por exemplo PETR4."),
    ("Evento", "Use Compra, Venda, Desdobramento, Grupamento, Amortização, Bonificação, Cisão, Resgate Antecipado ou Resgate Vencimento."),
    ("Data", "Informe a data do evento em DD/MM/AAAA, ou como data do Excel."),
    ("Quantidade", "Quantidade movimentada no evento."),
    ("Valor Evento", "Valor líquido em BRL."),
    ("Valor Bruto", "Campo cadastrado em vendas. Informe o valor bruto da venda em BRL."),
]

_INTERNATIONAL_TEMPLATE_INSTRUCTIONS = [
    ("Classe", "Use valores como Stock, REIT, ETF."),
    ("Ativo", "Ticker do ativo, por exemplo AAPL."),
    ("Evento", "Use Compra, Venda, Desdobramento, Grupamento, Amortização, Bonificação, Cisão, Resgate Antecipado ou Resgate Vencimento."),
    ("Data", "Informe a data do evento em DD/MM/AAAA, ou como data do Excel."),
    ("Quantidade", "Quantidade movimentada no evento."),
    ("Valor Evento", "Valor em USD."),
    ("Origem US", "Para compras no exterior antes de 2024, parcela paga com recursos já em USD."),
]


def build_import_template_xlsx(template: str) -> bytes:
    """Build an XLSX template compatible with the event import parser."""
    normalized = template.lower().strip()
    if normalized not in {"brasil", "exterior"}:
        raise ValueError("Modelo deve ser 'brasil' ou 'exterior'.")

    is_international = normalized == "exterior"
    headers = _INTERNATIONAL_TEMPLATE_HEADERS if is_international else _BR_TEMPLATE_HEADERS
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Exterior" if is_international else "Registro"
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = [18, 16, 22, 14, 14, 18, 18]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"

    instructions = workbook.create_sheet("Instruções")
    instructions.append(["Campo", "Orientação"])
    rows = _INTERNATIONAL_TEMPLATE_INSTRUCTIONS if is_international else _BR_TEMPLATE_INSTRUCTIONS
    for row in rows:
        instructions.append(row)
    for cell in instructions[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 90
    instructions.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────
# XLSX parser
# ─────────────────────────────────────────────────────────────

def _select_import_worksheet(wb) -> tuple:
    if "Registro" in wb.sheetnames:
        return wb["Registro"], "legacy"
    ws = wb.active
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    ]
    normalized = [h.lower() for h in headers]
    if "origem us" in normalized:
        return ws, "international"
    return ws, "legacy"


def parse_xlsx(source: Path | str | BinaryIO) -> list[dict]:
    """
    Parse an XLSX file and return a list of normalised event dicts
    ready for insertion.

    ``source`` can be a file path (Path/str) or an in-memory
    file-like object (e.g., BytesIO from an upload).

    Each dict contains:
        asset_class, ticker, event_type, event_date, quantity, event_value,
        gross_value, origin_usd

    Raises on unresolvable formulas or unknown labels.
    """
    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(str(source), data_only=False)
    else:
        wb = openpyxl.load_workbook(source, data_only=False)

    ws, layout = _select_import_worksheet(wb)

    events: list[dict] = []
    errors: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Legacy: A=Classe, B=Ativo, C=Evento, D=Data, E=Quantidade, F=Valor Evento, G=Valor Bruto
        # International: A=Classe, B=Ativo, C=Evento, D=Data, E=Quantidade, F=Valor Evento, G=Origem US
        raw_class = row[0].value
        raw_ticker = row[1].value
        raw_event = row[2].value
        raw_date = row[3].value
        raw_qty = row[4].value
        raw_value = row[5].value
        raw_extra_value = row[6].value if len(row) > 6 else None

        # Skip empty rows
        if not raw_class or not raw_ticker or not raw_event:
            continue

        # Map class
        class_str = str(raw_class).strip()
        asset_class = _CLASS_LABEL_MAP.get(class_str)
        if asset_class is None:
            errors.append(f"Linha {row_idx}: classe desconhecida '{class_str}'")
            continue

        # Map event type
        event_str = str(raw_event).strip()
        event_type = _EVENT_LABEL_MAP.get(event_str)
        if event_type is None:
            errors.append(f"Linha {row_idx}: evento desconhecido '{event_str}'")
            continue

        # Parse date
        if hasattr(raw_date, "strftime"):
            event_date = raw_date.strftime("%Y-%m-%d")
        elif isinstance(raw_date, str):
            event_date = raw_date.strip()
        else:
            errors.append(f"Linha {row_idx}: data inválida '{raw_date}'")
            continue

        # Parse quantity (may be formula)
        try:
            quantity = _resolve_cell_value(raw_qty, f"E{row_idx}")
        except (ValueError, InvalidOperation) as e:
            errors.append(f"Linha {row_idx}: erro na quantidade: {e}")
            continue

        # Parse value (may be formula)
        try:
            event_value = _resolve_cell_value(raw_value, f"F{row_idx}")
        except (ValueError, InvalidOperation) as e:
            errors.append(f"Linha {row_idx}: erro no valor: {e}")
            continue

        gross_value = None
        origin_usd = None
        if layout == "legacy" and event_type == EventType.VENDA and raw_extra_value not in (None, ""):
            try:
                gross_value = _resolve_cell_value(raw_extra_value, f"G{row_idx}")
            except (ValueError, InvalidOperation) as e:
                errors.append(f"Linha {row_idx}: erro no valor bruto: {e}")
                continue
        elif layout == "international" and event_type == EventType.COMPRA and raw_extra_value not in (None, ""):
            try:
                origin_usd = _resolve_cell_value(raw_extra_value, f"G{row_idx}")
            except (ValueError, InvalidOperation) as e:
                errors.append(f"Linha {row_idx}: erro na origem US: {e}")
                continue

        # For value-ignored events, force to zero
        if event_type in EventType.value_ignored():
            event_value = Decimal("0")
            gross_value = None
            origin_usd = None

        events.append({
            "asset_class": asset_class.value,
            "market": "US" if layout == "international" else None,
            "ticker": str(raw_ticker).strip().upper(),
            "event_type": event_type.value,
            "event_date": event_date,
            "quantity": str(quantity),
            "event_value": str(event_value),
            "gross_value": str(gross_value) if gross_value is not None else None,
            "origin_usd": str(origin_usd) if origin_usd is not None else None,
        })

    if errors:
        raise ImportError(
            f"{len(errors)} erro(s) na importação:\n" + "\n".join(errors)
        )

    return events


# ─────────────────────────────────────────────────────────────
# Duplicate detection
# ─────────────────────────────────────────────────────────────

def _check_duplicate(
    conn,
    portfolio_id: int,
    asset_id: int,
    event_type: str,
    event_date: str,
    quantity: str,
    event_value: str,
) -> int | None:
    """
    Check if an event with the same key fields already exists.

    Returns the existing event id if found, None otherwise.
    """
    row = conn.execute(
        """
        SELECT id FROM events
        WHERE portfolio_id = ?
          AND asset_id = ?
          AND event_type = ?
          AND event_date = ?
          AND quantity = ?
          AND event_value = ?
          AND is_cancelled = 0
          AND is_storno = 0
        LIMIT 1
        """,
        (portfolio_id, asset_id, event_type, event_date, quantity, event_value),
    ).fetchone()
    return row["id"] if row else None


def _flag_duplicate(conn, event_id: int, asset_id: int) -> None:
    """Set duplicate_flag on event and asset."""
    conn.execute(
        "UPDATE events SET duplicate_flag = 1 WHERE id = ?", (event_id,)
    )
    conn.execute(
        "UPDATE assets SET duplicate_flag = 1 WHERE id = ?", (asset_id,)
    )


# ─────────────────────────────────────────────────────────────
# Import pipeline
# ─────────────────────────────────────────────────────────────

def import_events_to_ledger(
    conn,
    parsed: list[dict],
    portfolio_id: int,
    source: str = "import_xlsx",
    notes_prefix: str = "Importado de planilha",
    source_row_offset: int = 1,
) -> dict:
    """
    Full import pipeline: parse XLSX → create assets → create events.

    Detects duplicate events and flags them instead of creating duplicates.

    Returns a summary dict with counts.
    """
    from backend.services.asset_service import match_asset, create_asset, build_operation_payload
    from backend.services.event_service import create_event, build_brl_conversion
    from backend.database import next_sequence

    imported = 0
    skipped = 0
    duplicates = 0
    duplicate_details: list[str] = []
    review_count = 0
    review_details: list[str] = []
    errors: list[str] = []

    for i, ev in enumerate(parsed, start=1):
        try:
            gross_value = ev.get("gross_value") if ev["event_type"] == EventType.VENDA.value else None
            origin_usd = ev.get("origin_usd") if ev["event_type"] == EventType.COMPRA.value else None
            operation_payload = build_operation_payload(
                ticker=ev["ticker"],
                asset_class=ev["asset_class"],
                market=ev.get("market"),
                event_date=ev["event_date"],
                portfolio_id=portfolio_id,
                event_type=ev["event_type"],
                quantity=ev["quantity"],
                event_value=ev["event_value"],
                gross_value=gross_value,
                origin_usd=origin_usd,
                notes=f"{notes_prefix} (linha {i + source_row_offset})",
                source_row=i + source_row_offset,
            )
            # Resolve or create asset through the central ticker + class + market matcher.
            match = match_asset(
                conn,
                ticker=ev["ticker"],
                asset_class=ev["asset_class"],
                event_date=ev["event_date"],
                market=ev.get("market"),
                source=source,
                create_review=True,
                operation_payload=operation_payload,
            )
            if match["status"] == "exact":
                asset_id = match["asset"]["id"]
            elif match["status"] == "none":
                asset = create_asset(
                    conn,
                    asset_class=ev["asset_class"],
                    ticker=ev["ticker"],
                    market=ev.get("market") or match.get("market"),
                    valid_from=None,
                    event_date=ev["event_date"],
                    source=source,
                )
                asset_id = asset["id"]
            else:
                review_count += 1
                skipped += 1
                review_details.append(
                    f"Linha {i + source_row_offset}: {ev['ticker']} ({ev['asset_class']}) em {ev['event_date']} enviado para revisão"
                )
                continue

            # Check for duplicates before insertion
            existing_id = _check_duplicate(
                conn,
                portfolio_id=portfolio_id,
                asset_id=asset_id,
                event_type=ev["event_type"],
                event_date=ev["event_date"],
                quantity=ev["quantity"],
                event_value=ev["event_value"],
            )

            if existing_id:
                # Insert the duplicate but with duplicate_flag = 1
                seq = next_sequence(conn)
                conversion = build_brl_conversion(
                    conn,
                    asset_id,
                    ev["event_date"],
                    to_decimal(ev["event_value"]),
                    to_decimal(gross_value) if gross_value is not None else None,
                )
                cur = conn.execute(
                    """
                    INSERT INTO events (portfolio_id, asset_id, event_type, event_date,
                                        quantity, event_value, event_value_brl, gross_value,
                                        gross_value_brl, ptax_compra, ptax_venda,
                                        sequence_num, duplicate_flag, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (
                        portfolio_id,
                        asset_id,
                        ev["event_type"],
                        ev["event_date"],
                        ev["quantity"],
                        ev["event_value"],
                        conversion["event_value_brl"],
                        gross_value,
                        conversion["gross_value_brl"],
                        conversion["ptax_compra"],
                        conversion["ptax_venda"],
                        seq,
                        "Possível duplicidade (Importação)",
                    ),
                )
                row = conn.execute("SELECT * FROM events WHERE id = ?", (cur.lastrowid,)).fetchone()
                if ev["event_type"] == EventType.COMPRA.value:
                    from backend.services.fiscal_lot_service import create_lot_for_purchase
                    create_lot_for_purchase(conn, row, origin_usd)
                from backend.services.event_service import recalculate_position
                recalculate_position(conn, asset_id, portfolio_id)
                
                # Flag asset
                conn.execute("UPDATE assets SET duplicate_flag = 1 WHERE id = ?", (asset_id,))
                
                imported += 1
                duplicates += 1
                duplicate_details.append(
                    f"Ativo {ev['ticker']}, {ev['event_type']} em {ev['event_date']}"
                )
            else:
                create_event(
                    conn,
                    portfolio_id=portfolio_id,
                    asset_id=asset_id,
                    event_type=ev["event_type"],
                    event_date=ev["event_date"],
                    quantity=ev["quantity"],
                    event_value=ev["event_value"],
                    gross_value=gross_value,
                    origin_usd=origin_usd,
                    notes=f"{notes_prefix} (linha {i + source_row_offset})",
                )
                imported += 1

        except Exception as e:
            errors.append(f"Evento {i} ({ev['ticker']} {ev['event_date']}): {e}")
            skipped += 1

    return {
        "total_rows": len(parsed),
        "imported": imported,
        "skipped": skipped,
        "duplicates": duplicates,
        "duplicate_details": duplicate_details,
        "review_count": review_count,
        "review_details": review_details,
        "errors": errors,
    }


def import_to_ledger(
    conn,
    source: Path | str | BinaryIO,
    portfolio_id: int,
) -> dict:
    """
    Full import pipeline: parse XLSX -> create assets -> create events.
    """
    parsed = parse_xlsx(source)
    return import_events_to_ledger(conn, parsed, portfolio_id)
