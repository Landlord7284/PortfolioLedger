"""B3 monthly XLSX importer.

The B3 spreadsheet is an external source. Market values and incomes are
persisted separately from the ledger; only B3 amortization income may create a
ledger event automatically when the asset/date/type key is not already present.
"""

from __future__ import annotations

import calendar
import hashlib
import json
import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Iterable

import openpyxl

from backend.domain.engine import to_decimal
from backend.domain.enums import AssetClass, B3IncomeEventStatus, B3MarketPriceStatus, B3MonthlyImportStatus, EventType
from backend.domain.normalization import (
    normalize_b3_income_event_status,
    normalize_b3_market_price_status,
    normalize_b3_monthly_import_status,
    normalize_bool_01,
    normalize_ticker,
)
from backend.services import asset_service, event_service


_FILENAME_RE = re.compile(r"^(?P<year>\d{4})-(?P<month>\d{2})\.xlsx$", re.IGNORECASE)
_TICKER_PREFIX_RE = re.compile(r"^\s*([A-Z]{4}\d{1,2}[A-Z]*)\s+-\s+(.+)$")
_TRADED_TICKER_RE = re.compile(r"^[A-Z]{4}\d{1,2}$")
_BR_TICKER_WITH_ALPHA_SUFFIX_RE = re.compile(r"^([A-Z]{4}\d{1,2})[A-Z]+$")
_FIXED_INCOME_CLASSES = {
    "DEB": AssetClass.DEBENTURE,
}
_SUMMARY_ONLY_PREFIXES = {"CRI", "CRA"}
_B3_REGISTRATION_COMPLETABLE_CLASSES = {
    AssetClass.ACAO.value,
    AssetClass.FII.value,
    AssetClass.FI_INFRA.value,
}


@dataclass
class SourceFile:
    filename: str
    content: bytes


@dataclass(frozen=True)
class ProductParse:
    original_ticker: str | None
    canonical_ticker: str | None
    extracted_name: str | None


def _norm_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch)).upper()


def _header_key(value) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", _norm_text(value)).strip()


def _clean_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text != "-" else None


def _clean_metadata_text(value) -> str | None:
    text = _clean_str(value)
    if text is None:
        return None
    return re.sub(r"\s+", " ", text).strip() or None


def _digits(value) -> str | None:
    text = _clean_str(value)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    return digits or None


def _decimal_str(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in {"", "-"}:
        return None
    return str(to_decimal(value))


def _clean_decimal_str(value) -> str | None:
    text = _clean_str(value)
    if text is None:
        return None
    cleaned = re.sub(r"[^0-9,.\-]", "", text.replace("'", "").replace('"', ""))
    if cleaned in {"", "-", ".", ","}:
        return None
    if "," not in cleaned and "." in cleaned:
        parts = cleaned.split(".")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]):
            cleaned = "".join(parts)
    return str(to_decimal(cleaned))


def _optional_ticker(value: str | None) -> str | None:
    return normalize_ticker(value) if value else None


def _json_dumps(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _json_loads(value: str | None) -> dict:
    if not value:
        return {}
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _date_str(value) -> str | None:
    if value is None or value == "-":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _month_from_filename(filename: str) -> tuple[str, str]:
    name = Path(filename).name
    match = _FILENAME_RE.match(name)
    if not match:
        raise ValueError(f"Arquivo B3 deve seguir o formato YYYY-MM.xlsx: {filename}")
    year = int(match.group("year"))
    month = int(match.group("month"))
    if month < 1 or month > 12:
        raise ValueError(f"Mes invalido no arquivo B3: {filename}")
    last_day = calendar.monthrange(year, month)[1]
    reference_month = f"{year:04d}-{month:02d}"
    reference_date = f"{reference_month}-{last_day:02d}"
    return reference_month, reference_date


def _worksheet_by_title(wb, wanted: str):
    wanted_key = _header_key(wanted)
    for name in wb.sheetnames:
        if _header_key(name) == wanted_key:
            return wb[name]
    return None


def _row_dicts(ws) -> Iterable[tuple[int, dict[str, object]]]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return
    headers = [_header_key(cell) for cell in header_row]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        if _clean_str(data.get("PRODUTO")) is None:
            break
        yield row_idx, data


def _extract_ticker_product(product: str | None) -> ProductParse:
    if not product:
        return ProductParse(None, None, None)
    match = _TICKER_PREFIX_RE.match(product.strip())
    if match:
        ticker = match.group(1).upper()
        suffix_match = _BR_TICKER_WITH_ALPHA_SUFFIX_RE.match(ticker)
        canonical_ticker = suffix_match.group(1) if suffix_match and _TRADED_TICKER_RE.match(suffix_match.group(1)) else None
        return ProductParse(ticker, canonical_ticker, _clean_metadata_text(match.group(2)))
    return ProductParse(None, None, _clean_metadata_text(product))


def _get_current_ticker(conn: sqlite3.Connection, asset_id: int) -> str | None:
    row = conn.execute(
        """
        SELECT ticker FROM asset_tickers
        WHERE asset_id = ? AND valid_until IS NULL
        ORDER BY valid_from DESC NULLS FIRST, id DESC LIMIT 1
        """,
        (asset_id,),
    ).fetchone()
    return row["ticker"] if row else None


def _single_asset(rows: list[sqlite3.Row]) -> dict | None:
    ids = []
    seen = set()
    for row in rows:
        asset_id = row["id"]
        if asset_id not in seen:
            ids.append(asset_id)
            seen.add(asset_id)
    return {"id": ids[0]} if len(ids) == 1 else None


def _find_by_ticker(conn: sqlite3.Connection, ticker: str | None, event_date: str | None) -> tuple[int | None, list[int]]:
    if not ticker:
        return None, []
    date_clause = ""
    params: list[object] = [normalize_ticker(ticker)]
    if event_date:
        date_clause = "AND (t.valid_from IS NULL OR t.valid_from <= ?) AND (t.valid_until IS NULL OR t.valid_until > ?)"
        params.extend([event_date, event_date])
    rows = conn.execute(
        f"""
        SELECT DISTINCT a.id
        FROM asset_tickers t
        JOIN assets a ON a.id = t.asset_id
        WHERE t.ticker = ?
          AND a.merged_into_asset_id IS NULL
          {date_clause}
        ORDER BY a.id
        """,
        params,
    ).fetchall()
    asset = _single_asset(rows)
    return (asset["id"], []) if asset else (None, [row["id"] for row in rows])


def _names_match(imported_name: str | None, candidate_names: list[str]) -> bool:
    names = [name for name in candidate_names if name]
    if not names:
        return True
    imported = _norm_text(imported_name)
    if not imported:
        return False
    imported_tokens = set(re.findall(r"[A-Z0-9]{3,}", imported))
    for name in names:
        candidate = _norm_text(name)
        if imported == candidate or imported in candidate or candidate in imported:
            return True
        candidate_tokens = set(re.findall(r"[A-Z0-9]{3,}", candidate))
        overlap = imported_tokens & candidate_tokens
        if len(overlap) >= 2 and len(overlap) * 10 >= max(1, min(len(imported_tokens), len(candidate_tokens))) * 6:
            return True
    return False


def _find_by_canonical_income_ticker(
    conn: sqlite3.Connection,
    ticker: str | None,
    imported_name: str | None,
    event_date: str | None,
) -> tuple[int | None, list[int]]:
    if not ticker:
        return None, []
    date_clause = ""
    params: list[object] = [normalize_ticker(ticker)]
    if event_date:
        date_clause = "AND (t.valid_from IS NULL OR t.valid_from <= ?) AND (t.valid_until IS NULL OR t.valid_until > ?)"
        params.extend([event_date, event_date])
    rows = conn.execute(
        f"""
        SELECT DISTINCT a.id, a.name AS asset_name, t.name AS ticker_name
        FROM asset_tickers t
        JOIN assets a ON a.id = t.asset_id
        WHERE t.ticker = ?
          AND a.merged_into_asset_id IS NULL
          {date_clause}
        ORDER BY a.id
        """,
        params,
    ).fetchall()
    asset = _single_asset(rows)
    candidates = [row["id"] for row in rows]
    if not asset:
        return None, candidates
    candidate_names = [row["asset_name"] for row in rows if row["id"] == asset["id"]]
    candidate_names.extend(row["ticker_name"] for row in rows if row["id"] == asset["id"])
    return (asset["id"], []) if _names_match(imported_name, candidate_names) else (None, [asset["id"]])


def _find_by_cnpj(conn: sqlite3.Connection, cnpj: str | None) -> tuple[int | None, list[int]]:
    if not cnpj:
        return None, []
    rows = conn.execute(
        """
        SELECT id FROM assets
        WHERE REPLACE(REPLACE(REPLACE(COALESCE(cnpj, ''), '.', ''), '/', ''), '-', '') = ?
          AND merged_into_asset_id IS NULL
        ORDER BY id
        """,
        (cnpj,),
    ).fetchall()
    asset = _single_asset(rows)
    return (asset["id"], []) if asset else (None, [row["id"] for row in rows])


def _find_by_name(conn: sqlite3.Connection, name: str | None, asset_class: str | None = None) -> tuple[int | None, list[int]]:
    if not name:
        return None, []
    normalized = _norm_text(name)
    rows = conn.execute(
        """
        SELECT a.*, t.name AS ticker_name
        FROM assets a
        LEFT JOIN asset_tickers t ON t.asset_id = a.id AND t.valid_until IS NULL
        WHERE a.merged_into_asset_id IS NULL
        ORDER BY a.id
        """
    ).fetchall()
    matches = []
    for row in rows:
        if asset_class and row["asset_class"] != asset_class:
            continue
        candidates = [row["name"], row["ticker_name"]]
        if any(candidate and (_norm_text(candidate) == normalized or normalized in _norm_text(candidate) or _norm_text(candidate) in normalized) for candidate in candidates):
            matches.append(row)
    asset = _single_asset(matches)
    return (asset["id"], []) if asset else (None, [row["id"] for row in matches])


def _find_tesouro(conn: sqlite3.Connection, product: str | None, maturity_date: str | None) -> tuple[int | None, list[int]]:
    asset_id, candidates = _find_by_name(conn, product, AssetClass.TESOURO_DIRETO.value)
    if asset_id or candidates:
        return asset_id, candidates
    if not maturity_date:
        return None, []
    rows = conn.execute(
        """
        SELECT id FROM assets
        WHERE asset_class = ? AND maturity_date = ? AND merged_into_asset_id IS NULL
        ORDER BY id
        """,
        (AssetClass.TESOURO_DIRETO.value, maturity_date),
    ).fetchall()
    asset = _single_asset(rows)
    return (asset["id"], []) if asset else (None, [row["id"] for row in rows])


def _create_review(
    conn: sqlite3.Connection,
    ticker: str | None,
    asset_class: str,
    event_date: str | None,
    candidates: list[int],
    reason: str,
    payload: dict,
) -> int:
    review = asset_service.create_match_review(
        conn,
        source="b3_monthly_import",
        ticker=ticker or payload.get("product") or "B3",
        asset_class=asset_class,
        market="BR",
        event_date=event_date,
        candidate_asset_ids=candidates,
        reason=reason,
        operation_payload=payload,
    )
    return review["id"]


def _income_raw_field(raw_payload: dict, key: str) -> str | None:
    return _clean_str(raw_payload.get(_header_key(key)))


def _income_fingerprint(
    *,
    portfolio_id: int,
    reference_month: str,
    product: str | None,
    ticker: str | None,
    payment_date: str,
    event_type: str,
    quantity: str | None,
    unit_price: str | None,
    net_value: str,
    institution: str | None,
    account: str | None,
) -> str:
    parts = [
        str(portfolio_id),
        reference_month,
        _norm_text(product),
        payment_date,
        _norm_text(event_type),
        quantity or "",
        unit_price or "",
        net_value,
        _norm_text(institution),
        _norm_text(account),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _income_fingerprint_from_values(values: dict, reference_month: str) -> str:
    raw_payload = values.get("raw_payload") if isinstance(values.get("raw_payload"), dict) else {}
    return _income_fingerprint(
        portfolio_id=values["portfolio_id"],
        reference_month=reference_month,
        product=values.get("product"),
        ticker=values.get("ticker"),
        payment_date=values["payment_date"],
        event_type=values["event_type"],
        quantity=values.get("quantity"),
        unit_price=values.get("unit_price"),
        net_value=values.get("net_value") or "0",
        institution=_income_raw_field(raw_payload, "Instituição"),
        account=_income_raw_field(raw_payload, "Conta"),
    )


def _income_fingerprint_from_row(row: sqlite3.Row, reference_month: str) -> str:
    raw_payload = _json_loads(row["raw_payload"])
    return _income_fingerprint(
        portfolio_id=row["portfolio_id"],
        reference_month=reference_month,
        product=row["product"],
        ticker=row["ticker"],
        payment_date=row["payment_date"],
        event_type=row["event_type"],
        quantity=row["quantity"],
        unit_price=row["unit_price"],
        net_value=row["net_value"] or "0",
        institution=_income_raw_field(raw_payload, "Instituição"),
        account=_income_raw_field(raw_payload, "Conta"),
    )


def _find_income_decision(conn: sqlite3.Connection, values: dict, reference_month: str) -> sqlite3.Row | None:
    fingerprint = _income_fingerprint_from_values(values, reference_month)
    return conn.execute(
        """
        SELECT d.*
        FROM b3_income_resolution_decisions d
        JOIN assets a ON a.id = d.resolved_asset_id
        WHERE d.portfolio_id = ?
          AND d.fingerprint = ?
          AND a.merged_into_asset_id IS NULL
        """,
        (values["portfolio_id"], fingerprint),
    ).fetchone()


def _upsert_income_decision(conn: sqlite3.Connection, income: sqlite3.Row, asset_id: int) -> None:
    import_row = conn.execute(
        "SELECT reference_month FROM b3_monthly_imports WHERE id = ?",
        (income["import_id"],),
    ).fetchone()
    if not import_row:
        raise ValueError("Import B3 de origem nao encontrado.")
    reference_month = import_row["reference_month"]
    raw_payload = _json_loads(income["raw_payload"])
    institution = _income_raw_field(raw_payload, "Instituição")
    account = _income_raw_field(raw_payload, "Conta")
    fingerprint = _income_fingerprint_from_row(income, reference_month)
    conn.execute(
        """
        INSERT INTO b3_income_resolution_decisions (
            portfolio_id, reference_month, source_sheet, source_row, product, ticker,
            payment_date, event_type, net_value, institution, account, fingerprint,
            resolved_asset_id
        )
        VALUES (?, ?, 'Proventos Recebidos', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(portfolio_id, fingerprint) DO UPDATE SET
            reference_month = excluded.reference_month,
            source_row = excluded.source_row,
            product = excluded.product,
            ticker = excluded.ticker,
            payment_date = excluded.payment_date,
            event_type = excluded.event_type,
            net_value = excluded.net_value,
            institution = excluded.institution,
            account = excluded.account,
            resolved_asset_id = excluded.resolved_asset_id,
            updated_at = datetime('now')
        """,
        (
            income["portfolio_id"],
            reference_month,
            income["source_row"],
            income["product"],
            income["ticker"],
            income["payment_date"],
            income["event_type"],
            income["net_value"] or "0",
            institution,
            account,
            fingerprint,
            asset_id,
        ),
    )


def _resolve_asset(
    conn: sqlite3.Connection,
    *,
    ticker: str | None,
    product: str | None,
    cnpj: str | None,
    maturity_date: str | None,
    asset_class: AssetClass,
    event_date: str | None,
) -> tuple[int | None, list[int], str | None]:
    asset_id, candidates = _find_by_ticker(conn, ticker, event_date)
    if asset_id or candidates:
        return asset_id, candidates, "ticker"
    asset_id, candidates = _find_by_cnpj(conn, cnpj)
    if asset_id or candidates:
        return asset_id, candidates, "cnpj"
    if asset_class == AssetClass.TESOURO_DIRETO:
        asset_id, candidates = _find_tesouro(conn, product, maturity_date)
        return asset_id, candidates, "tesouro"
    asset_id, candidates = _find_by_name(conn, product, asset_class.value)
    return asset_id, candidates, "name"


def _product_registration_name(product: str | None, ticker: str | None) -> str | None:
    parsed = _extract_ticker_product(product)
    if not parsed.original_ticker or not parsed.extracted_name:
        return None
    if ticker and parsed.original_ticker != normalize_ticker(ticker):
        return None
    return parsed.extracted_name


def _complete_existing_b3_asset_metadata(
    conn: sqlite3.Connection,
    asset_id: int | None,
    *,
    product: str | None,
    ticker: str | None,
    issuer_name: str | None,
    cnpj: str | None,
    maturity_date: str | None,
) -> None:
    if not asset_id:
        return
    asset = asset_service.get_asset(conn, asset_id)
    if not asset:
        return

    updates = {}
    if asset["asset_class"] == AssetClass.DEBENTURE.value:
        imported_name = _clean_metadata_text(issuer_name) or _clean_metadata_text(product)
        if _clean_str(asset.get("name")) is None and imported_name is not None:
            updates["name"] = imported_name
        if _clean_str(asset.get("maturity_date")) is None and maturity_date is not None:
            updates["maturity_date"] = maturity_date
    elif asset["asset_class"] in _B3_REGISTRATION_COMPLETABLE_CLASSES:
        imported_name = _product_registration_name(product, ticker)
        if _clean_str(asset.get("name")) is None and imported_name is not None:
            updates["name"] = imported_name
        if _clean_str(asset.get("cnpj")) is None and cnpj is not None:
            updates["cnpj"] = cnpj

    if not updates:
        return

    asset_service.update_asset_metadata(
        conn,
        asset_id,
        **updates,
    )


def _upsert_import(conn: sqlite3.Connection, portfolio_id: int, filename: str, reference_month: str, reference_date: str) -> int:
    processed_status = normalize_b3_monthly_import_status(B3MonthlyImportStatus.PROCESSED.value)
    conn.execute(
        f"""
        INSERT INTO b3_monthly_imports (portfolio_id, filename, reference_month, reference_date)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(portfolio_id, filename) DO UPDATE SET
            reference_month = excluded.reference_month,
            reference_date = excluded.reference_date,
            status = '{processed_status}',
            total_rows = 0,
            imported_prices = 0,
            imported_incomes = 0,
            auto_events_created = 0,
            duplicates = 0,
            review_count = 0,
            errors = NULL,
            updated_at = datetime('now')
        """,
        (portfolio_id, filename, reference_month, reference_date),
    )
    row = conn.execute(
        "SELECT id FROM b3_monthly_imports WHERE portfolio_id = ? AND filename = ?",
        (portfolio_id, filename),
    ).fetchone()
    return row["id"]


def _upsert_market_price(conn: sqlite3.Connection, values: dict) -> str:
    status = normalize_b3_market_price_status(values["status"])
    is_unit_price = normalize_bool_01(values["is_unit_price"])
    ticker = _optional_ticker(values.get("ticker"))
    existing = conn.execute(
        """
        SELECT * FROM b3_market_prices
        WHERE import_id = ? AND source_sheet = ? AND source_row = ?
        """,
        (values["import_id"], values["source_sheet"], values["source_row"]),
    ).fetchone()
    if existing:
        if not is_unit_price:
            try:
                existing_payload = json.loads(existing["raw_payload"] or "{}")
            except json.JSONDecodeError:
                existing_payload = {}
            if len(existing_payload.get("consolidated_source_rows") or []) > 1:
                return "consolidated_reprocessed"
        conn.execute(
            """
            UPDATE b3_market_prices
            SET asset_id = ?,
                reference_month = ?,
                reference_date = ?,
                ticker = ?,
                product = ?,
                cnpj = ?,
                maturity_date = ?,
                value = ?,
                is_unit_price = ?,
                status = ?,
                review_id = ?,
                raw_payload = ?,
                updated_at = datetime('now')
            WHERE id = ?
            """,
            (
                values["asset_id"],
                values["reference_month"],
                values["reference_date"],
                ticker,
                values.get("product"),
                values.get("cnpj"),
                values.get("maturity_date"),
                values.get("value"),
                is_unit_price,
                status,
                values.get("review_id"),
                _json_dumps(values["raw_payload"]),
                existing["id"],
            ),
        )
        return "updated"

    if not is_unit_price and values.get("asset_id") is not None:
        asset_existing = conn.execute(
            """
            SELECT * FROM b3_market_prices
            WHERE asset_id = ?
              AND reference_month = ?
              AND source_sheet = ?
            ORDER BY id
            LIMIT 1
            """,
            (values["asset_id"], values["reference_month"], values["source_sheet"]),
        ).fetchone()
        if asset_existing:
            try:
                existing_payload = json.loads(asset_existing["raw_payload"] or "{}")
            except json.JSONDecodeError:
                existing_payload = {}
            source_rows = existing_payload.get("consolidated_source_rows") or [asset_existing["source_row"]]
            if values["source_row"] in source_rows:
                return "consolidated_reprocessed"
            current_value = to_decimal(asset_existing["value"]) if asset_existing["value"] is not None else Decimal("0")
            added_value = to_decimal(values["value"]) if values.get("value") is not None else Decimal("0")
            source_rows.append(values["source_row"])
            conn.execute(
                """
                UPDATE b3_market_prices
                SET value = ?,
                    ticker = ?,
                    product = ?,
                    cnpj = ?,
                    maturity_date = ?,
                    status = ?,
                    review_id = ?,
                    raw_payload = ?,
                    updated_at = datetime('now')
                WHERE id = ?
                """,
                (
                    str(current_value + added_value),
                    ticker,
                    values.get("product"),
                    values.get("cnpj"),
                    values.get("maturity_date"),
                    status,
                    values.get("review_id"),
                    _json_dumps(
                        {
                            "consolidated_source_rows": source_rows,
                            "latest_row": values["raw_payload"],
                        }
                    ),
                    asset_existing["id"],
                ),
            )
            return "consolidated"

    conn.execute(
        """
        INSERT INTO b3_market_prices (
            import_id, asset_id, reference_month, reference_date, source_sheet,
            source_row, ticker, product, cnpj, maturity_date, value,
            is_unit_price, status, review_id, raw_payload
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            values["import_id"],
            values["asset_id"],
            values["reference_month"],
            values["reference_date"],
            values["source_sheet"],
            values["source_row"],
            ticker,
            values.get("product"),
            values.get("cnpj"),
            values.get("maturity_date"),
            values.get("value"),
            is_unit_price,
            status,
            values.get("review_id"),
            _json_dumps(values["raw_payload"]),
        ),
    )
    return "inserted"


def _upsert_income(conn: sqlite3.Connection, values: dict) -> bool:
    status = normalize_b3_income_event_status(values["status"])
    ticker = _optional_ticker(values.get("ticker"))
    existing = conn.execute(
        """
        SELECT * FROM b3_income_events
        WHERE import_id = ? AND source_row = ?
        """,
        (values["import_id"], values["source_row"]),
    ).fetchone()
    if existing:
        if existing["status"] == B3IncomeEventStatus.DISCARDED.value:
            return False
        conn.execute(
            """
            UPDATE b3_income_events
            SET asset_id = ?,
                payment_date = ?,
                event_type = ?,
                product = ?,
                ticker = ?,
                quantity = ?,
                unit_price = ?,
                net_value = ?,
                status = ?,
                ledger_event_id = COALESCE(ledger_event_id, ?),
                review_id = ?,
                raw_payload = ?,
                updated_at = datetime('now')
            WHERE id = ?
            """,
            (
                values["asset_id"],
                values["payment_date"],
                values["event_type"],
                values["product"],
                ticker,
                values.get("quantity"),
                values.get("unit_price"),
                values.get("net_value"),
                status,
                values.get("ledger_event_id"),
                values.get("review_id"),
                _json_dumps(values["raw_payload"]),
                existing["id"],
            ),
        )
        return False
    conn.execute(
        """
        INSERT INTO b3_income_events (
            import_id, portfolio_id, asset_id, source_row, payment_date, event_type,
            product, ticker, quantity, unit_price, net_value, status,
            ledger_event_id, review_id, raw_payload
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            values["import_id"],
            values["portfolio_id"],
            values["asset_id"],
            values["source_row"],
            values["payment_date"],
            values["event_type"],
            values["product"],
            ticker,
            values.get("quantity"),
            values.get("unit_price"),
            values.get("net_value"),
            status,
            values.get("ledger_event_id"),
            values.get("review_id"),
            _json_dumps(values["raw_payload"]),
        ),
    )
    return True


def _existing_income(conn: sqlite3.Connection, import_id: int, source_row: int) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT * FROM b3_income_events
        WHERE import_id = ? AND source_row = ?
        """,
        (import_id, source_row),
    ).fetchone()


def _income_pending_status(row: sqlite3.Row) -> str:
    if row["income_status"] == B3IncomeEventStatus.DISCARDED.value:
        return "discarded"
    if row["income_status"] == B3IncomeEventStatus.REVIEW.value:
        return "pending"
    return "resolved"


def _income_pending_response(row: sqlite3.Row) -> dict:
    raw_payload = _json_loads(row["raw_payload"])
    return {
        "id": row["id"],
        "import_id": row["import_id"],
        "portfolio_id": row["portfolio_id"],
        "reference_month": row["reference_month"],
        "filename": row["filename"],
        "source_row": row["source_row"],
        "payment_date": row["payment_date"],
        "event_type": row["event_type"],
        "product": row["product"],
        "ticker": row["ticker"],
        "quantity": row["quantity"],
        "unit_price": row["unit_price"],
        "net_value": row["net_value"],
        "currency": "BRL",
        "institution": _income_raw_field(raw_payload, "Instituição"),
        "account": _income_raw_field(raw_payload, "Conta"),
        "status": _income_pending_status(row),
        "reason": row["reason"],
        "review_id": row["review_id"],
        "candidate_asset_ids": row["candidate_asset_ids"],
        "asset_id": row["asset_id"],
        "ledger_event_id": row["ledger_event_id"],
        "raw_payload": row["raw_payload"],
    }


def list_income_pendings(
    conn: sqlite3.Connection,
    portfolio_id: int,
    status: str = "pending",
) -> list[dict]:
    conditions = ["i.portfolio_id = ?"]
    params: list[object] = [portfolio_id]
    if status == "pending":
        conditions.append("i.status = ?")
        params.append(B3IncomeEventStatus.REVIEW.value)
    elif status == "resolved":
        conditions.append("i.status NOT IN (?, ?)")
        params.extend([B3IncomeEventStatus.REVIEW.value, B3IncomeEventStatus.DISCARDED.value])
        conditions.append("i.asset_id IS NOT NULL")
    elif status == "discarded":
        conditions.append("i.status = ?")
        params.append(B3IncomeEventStatus.DISCARDED.value)
    elif status != "all":
        raise ValueError("Status de pendencia B3 invalido.")
    rows = conn.execute(
        f"""
        SELECT
            i.id,
            i.import_id,
            i.portfolio_id,
            i.asset_id,
            i.source_row,
            i.payment_date,
            i.event_type,
            i.product,
            i.ticker,
            i.quantity,
            i.unit_price,
            i.net_value,
            i.status AS income_status,
            i.ledger_event_id,
            i.review_id,
            i.raw_payload,
            m.reference_month,
            m.filename,
            r.reason,
            r.candidate_asset_ids
        FROM b3_income_events i
        JOIN b3_monthly_imports m ON m.id = i.import_id
        LEFT JOIN asset_match_reviews r ON r.id = i.review_id
        WHERE {' AND '.join(conditions)}
        ORDER BY m.reference_month DESC, i.payment_date DESC, i.id DESC
        """,
        params,
    ).fetchall()
    return [_income_pending_response(row) for row in rows]


def _get_income_for_decision(conn: sqlite3.Connection, income_id: int, portfolio_id: int | None = None) -> sqlite3.Row | None:
    conditions = ["i.id = ?"]
    params: list[object] = [income_id]
    if portfolio_id is not None:
        conditions.append("i.portfolio_id = ?")
        params.append(portfolio_id)
    return conn.execute(
        f"""
        SELECT i.*, m.reference_month, m.filename, r.reason, r.candidate_asset_ids,
               i.status AS income_status
        FROM b3_income_events i
        JOIN b3_monthly_imports m ON m.id = i.import_id
        LEFT JOIN asset_match_reviews r ON r.id = i.review_id
        WHERE {' AND '.join(conditions)}
        """,
        params,
    ).fetchone()


def resolve_income_pending(conn: sqlite3.Connection, income_id: int, asset_id: int) -> dict:
    income = _get_income_for_decision(conn, income_id)
    if not income:
        raise ValueError("Pendencia B3 nao encontrada.")
    if income["income_status"] == B3IncomeEventStatus.DISCARDED.value:
        raise ValueError("Pendencia B3 descartada nao pode ser resolvida.")
    if income["income_status"] != B3IncomeEventStatus.REVIEW.value:
        raise ValueError("Pendencia B3 ja foi resolvida.")
    asset = asset_service.get_asset(conn, asset_id)
    if not asset or asset.get("merged_into_asset_id"):
        raise ValueError("Ativo destino nao encontrado ou esta mesclado.")

    ledger_event_id = income["ledger_event_id"]
    status = B3IncomeEventStatus.IMPORTED.value
    if _is_amortization(income["event_type"]):
        if ledger_event_id:
            raise ValueError("Pendencia B3 ja possui evento ledger vinculado.")
        ledger_event_id, _duplicate_reason = _auto_create_amortization(
            conn,
            income["portfolio_id"],
            asset_id,
            income["payment_date"],
            income["net_value"] or "0",
        )
        status = B3IncomeEventStatus.LEDGER_EVENT_CREATED.value

    ticker = _optional_ticker(_get_current_ticker(conn, asset_id) or income["ticker"])
    conn.execute(
        """
        UPDATE b3_income_events
        SET asset_id = ?,
            ticker = ?,
            status = ?,
            ledger_event_id = COALESCE(ledger_event_id, ?),
            updated_at = datetime('now')
        WHERE id = ?
        """,
        (asset_id, ticker, status, ledger_event_id, income_id),
    )
    if income["review_id"]:
        asset_service.resolve_match_review(conn, income["review_id"])
    updated = _get_income_for_decision(conn, income_id)
    _upsert_income_decision(conn, updated, asset_id)
    return _income_pending_response(updated)


def discard_income_pending(conn: sqlite3.Connection, income_id: int) -> dict:
    income = _get_income_for_decision(conn, income_id)
    if not income:
        raise ValueError("Pendencia B3 nao encontrada.")
    if income["ledger_event_id"]:
        raise ValueError("Pendencia B3 com evento ledger vinculado nao pode ser descartada.")
    if income["income_status"] != B3IncomeEventStatus.REVIEW.value:
        raise ValueError("Apenas pendencias B3 em aberto podem ser descartadas.")
    conn.execute(
        """
        UPDATE b3_income_events
        SET status = ?, updated_at = datetime('now')
        WHERE id = ?
        """,
        (B3IncomeEventStatus.DISCARDED.value, income_id),
    )
    if income["review_id"]:
        asset_service.resolve_match_review(conn, income["review_id"])
    updated = _get_income_for_decision(conn, income_id)
    return _income_pending_response(updated)


def _ledger_amortization_exists(conn: sqlite3.Connection, portfolio_id: int, asset_id: int, payment_date: str) -> bool:
    row = conn.execute(
        """
        SELECT 1 FROM events
        WHERE portfolio_id = ?
          AND asset_id = ?
          AND event_type = ?
          AND event_date = ?
          AND is_cancelled = 0
          AND is_storno = 0
        LIMIT 1
        """,
        (portfolio_id, asset_id, EventType.AMORTIZACAO.value, payment_date),
    ).fetchone()
    return row is not None


def _is_amortization(label: str) -> bool:
    return _norm_text(label) == _norm_text(EventType.AMORTIZACAO.value)


def _auto_create_amortization(conn: sqlite3.Connection, portfolio_id: int, asset_id: int, payment_date: str, net_value: str) -> tuple[int | None, str | None]:
    duplicate = _ledger_amortization_exists(conn, portfolio_id, asset_id, payment_date)
    event = event_service.create_event(
        conn,
        portfolio_id=portfolio_id,
        asset_id=asset_id,
        event_type=EventType.AMORTIZACAO.value,
        event_date=payment_date,
        quantity="0",
        event_value=net_value,
        notes="Importado automaticamente de Proventos Recebidos B3",
    )
    if duplicate:
        conn.execute("UPDATE events SET duplicate_flag = 1 WHERE id = ?", (event["id"],))
        conn.execute("UPDATE assets SET duplicate_flag = 1 WHERE id = ?", (asset_id,))
        return event["id"], "duplicate"
    return event["id"], None


def _process_market_sheet(
    conn: sqlite3.Connection,
    ws,
    *,
    import_id: int,
    reference_month: str,
    reference_date: str,
    sheet_name: str,
    asset_class: AssetClass,
    value_header: str,
    is_unit_price: bool,
    product_header: str = "PRODUTO",
    ticker_header: str | None = None,
    cnpj_header: str | None = None,
    maturity_header: str | None = None,
    fixed_income_only: bool = False,
    fixed_income_positions: list[dict] | None = None,
) -> dict:
    result = {
        "total_rows": 0,
        "imported_prices": 0,
        "duplicates": 0,
        "duplicate_details": [],
        "review_count": 0,
        "review_details": [],
        "errors": [],
    }
    for row_idx, row in _row_dicts(ws):
        product = _clean_str(row.get(product_header))
        if _norm_text(product).startswith("TOTAL"):
            continue
        if fixed_income_only:
            prefix = _norm_text(product).split(" - ")[0].strip()
            if prefix not in _FIXED_INCOME_CLASSES:
                continue
            asset_class = _FIXED_INCOME_CLASSES[prefix]
        result["total_rows"] += 1
        ticker = _optional_ticker(_clean_str(row.get(ticker_header))) if ticker_header else None
        cnpj = _digits(row.get(cnpj_header)) if cnpj_header else None
        maturity_date = _date_str(row.get(maturity_header)) if maturity_header else None
        quantity = None
        try:
            quantity = _clean_decimal_str(row.get("QUANTIDADE"))
        except (TypeError, ValueError):
            quantity = None
        try:
            value = _decimal_str(row.get(value_header))
        except (TypeError, ValueError) as exc:
            result["errors"].append(f"{sheet_name} linha {row_idx}: valor de mercado invalido: {exc}")
            continue
        if value is None:
            result["errors"].append(f"{sheet_name} linha {row_idx}: valor de mercado ausente ou invalido")
            continue
        asset_id, candidates, _source = _resolve_asset(
            conn,
            ticker=ticker,
            product=product,
            cnpj=cnpj,
            maturity_date=maturity_date,
            asset_class=asset_class,
            event_date=reference_date,
        )
        _complete_existing_b3_asset_metadata(
            conn,
            asset_id,
            product=product,
            ticker=ticker,
            issuer_name=_clean_str(row.get("EMISSOR")),
            cnpj=cnpj,
            maturity_date=maturity_date,
        )
        review_id = None
        status = B3MarketPriceStatus.IMPORTED.value
        if not asset_id:
            if fixed_income_only and asset_class == AssetClass.DEBENTURE and ticker and not candidates:
                asset = asset_service.create_asset(
                    conn,
                    AssetClass.DEBENTURE.value,
                    ticker,
                    market="BR",
                    name=_clean_str(row.get("EMISSOR")) or product,
                    maturity_date=maturity_date,
                    source="b3_monthly_import",
                )
                asset_id = asset["id"]
            else:
                payload = {
                    "sheet": sheet_name,
                    "source_row": row_idx,
                    "ticker": ticker,
                    "product": product,
                    "cnpj": cnpj,
                    "maturity_date": maturity_date,
                    "reference_date": reference_date,
                }
                review_id = _create_review(conn, ticker, asset_class.value, reference_date, candidates, "Ativo B3 nao resolvido com seguranca.", payload)
                result["review_count"] += 1
                result["review_details"].append(f"{sheet_name} linha {row_idx}: {ticker or product} enviado para revisao")
                status = B3MarketPriceStatus.REVIEW.value
        if fixed_income_only and fixed_income_positions is not None and asset_class == AssetClass.DEBENTURE and ticker and quantity:
            fixed_income_positions.append(
                {
                    "product_key": _norm_text(product),
                    "quantity": quantity,
                    "ticker": ticker,
                    "asset_id": asset_id,
                    "product": product,
                }
            )
        upsert_status = _upsert_market_price(
            conn,
            {
                "import_id": import_id,
                "asset_id": asset_id,
                "reference_month": reference_month,
                "reference_date": reference_date,
                "source_sheet": sheet_name,
                "source_row": row_idx,
                "ticker": ticker,
                "product": product,
                "cnpj": cnpj,
                "maturity_date": maturity_date,
                "value": value,
                "is_unit_price": is_unit_price,
                "status": status,
                "review_id": review_id,
                "raw_payload": row,
            },
        )
        if upsert_status in {"inserted", "consolidated"}:
            result["imported_prices"] += 1
        else:
            result["duplicates"] += 1
            duplicate_reason = (
                "posicao consolidada ja reprocessada"
                if upsert_status == "consolidated_reprocessed"
                else "linha ja reprocessada"
            )
            result["duplicate_details"].append(
                f"{sheet_name} linha {row_idx}: {ticker or product} {duplicate_reason} para {reference_month}"
            )
    return result


def _infer_income_class(ticker: str | None, product: str | None) -> AssetClass:
    prefix = _norm_text(product).split(" - ")[0].strip()
    if prefix == "DEB":
        return AssetClass.DEBENTURE
    if prefix == "CRI":
        return AssetClass.CRI
    if prefix == "CRA":
        return AssetClass.CRA
    if not ticker and _norm_text(product).startswith("TESOURO "):
        return AssetClass.TESOURO_DIRETO
    if ticker and ticker.endswith("11"):
        return AssetClass.FII
    return AssetClass.ACAO


def _resolve_income_asset(
    conn: sqlite3.Connection,
    *,
    parsed_product: ProductParse,
    product: str,
    asset_class: AssetClass,
    event_date: str,
) -> tuple[int | None, list[int], str | None, str | None]:
    asset_id, candidates = _find_by_ticker(conn, parsed_product.original_ticker, event_date)
    if asset_id or candidates:
        return asset_id, candidates, "ticker", parsed_product.original_ticker
    asset_id, candidates = _find_by_canonical_income_ticker(
        conn,
        parsed_product.canonical_ticker,
        parsed_product.extracted_name,
        event_date,
    )
    if asset_id or candidates:
        return asset_id, candidates, "canonical_ticker", parsed_product.canonical_ticker
    if asset_class == AssetClass.TESOURO_DIRETO:
        asset_id, candidates = _find_tesouro(conn, parsed_product.extracted_name or product, None)
        return asset_id, candidates, "tesouro", parsed_product.original_ticker
    asset_id, candidates = _find_by_name(conn, parsed_product.extracted_name or product, asset_class.value)
    return asset_id, candidates, "name", parsed_product.original_ticker


def _match_fixed_income_income(product: str | None, quantity: str | None, fixed_income_positions: list[dict]) -> dict | None:
    if not product or not quantity:
        return None
    product_key = _norm_text(product)
    matches = [
        item for item in fixed_income_positions
        if item["product_key"] == product_key and item["quantity"] == quantity
    ]
    if len(matches) == 1:
        return matches[0]
    return None


def _process_income_sheet(
    conn: sqlite3.Connection,
    ws,
    *,
    import_id: int,
    portfolio_id: int,
    reference_month: str,
    fixed_income_positions: list[dict] | None = None,
) -> dict:
    fixed_income_positions = fixed_income_positions or []
    result = {
        "total_rows": 0,
        "imported_incomes": 0,
        "auto_events_created": 0,
        "duplicates": 0,
        "duplicate_details": [],
        "review_count": 0,
        "review_details": [],
        "errors": [],
    }
    for row_idx, row in _row_dicts(ws):
        result["total_rows"] += 1
        product = _clean_str(row.get("PRODUTO"))
        if _norm_text(product).startswith("TOTAL"):
            continue
        parsed_product = _extract_ticker_product(product)
        ticker = _optional_ticker(parsed_product.canonical_ticker or parsed_product.original_ticker)
        payment_date = _date_str(row.get("PAGAMENTO"))
        event_type = _clean_str(row.get("TIPO DE EVENTO"))
        try:
            quantity = _clean_decimal_str(row.get("QUANTIDADE"))
            unit_price = _decimal_str(row.get("PRECO UNITARIO"))
            net_value = _decimal_str(row.get("VALOR LIQUIDO"))
        except (TypeError, ValueError) as exc:
            result["errors"].append(f"Proventos Recebidos linha {row_idx}: valor invalido: {exc}")
            continue
        if not product or not payment_date or not event_type or net_value is None:
            result["errors"].append(f"Proventos Recebidos linha {row_idx}: campos obrigatorios ausentes")
            continue
        inferred_class = _infer_income_class(ticker, product)
        fixed_income_match = _match_fixed_income_income(product, quantity, fixed_income_positions) if inferred_class == AssetClass.DEBENTURE else None
        if fixed_income_match:
            ticker = _optional_ticker(fixed_income_match["ticker"])
            asset_id = fixed_income_match["asset_id"]
            candidates = []
        else:
            asset_id, candidates, _source, ticker = _resolve_income_asset(
                conn,
                parsed_product=parsed_product,
                product=product,
                asset_class=inferred_class,
                event_date=payment_date,
            )
            ticker = _optional_ticker(ticker)
        review_id = None
        status = B3IncomeEventStatus.IMPORTED.value
        ledger_event_id = None
        income_values = {
            "import_id": import_id,
            "portfolio_id": portfolio_id,
            "asset_id": asset_id,
            "source_row": row_idx,
            "payment_date": payment_date,
            "event_type": event_type,
            "product": product,
            "ticker": ticker,
            "quantity": quantity,
            "unit_price": unit_price,
            "net_value": net_value,
            "raw_payload": row,
        }
        prefix = _norm_text(product).split(" - ")[0].strip()
        summary_only = (
            prefix in _SUMMARY_ONLY_PREFIXES
            or (not ticker and inferred_class not in {AssetClass.DEBENTURE, AssetClass.TESOURO_DIRETO})
        )
        if not asset_id:
            decision = _find_income_decision(conn, income_values, reference_month)
            if decision:
                asset_id = decision["resolved_asset_id"]
                income_values["asset_id"] = asset_id
                ticker = _optional_ticker(_get_current_ticker(conn, asset_id) or ticker)
                income_values["ticker"] = ticker
        if not asset_id and summary_only:
            status = B3IncomeEventStatus.SUMMARY_ONLY.value
        elif not asset_id:
            payload = {
                "sheet": "Proventos Recebidos",
                "source_row": row_idx,
                "ticker": ticker,
                "product": product,
                "payment_date": payment_date,
                "event_type": event_type,
                "quantity": quantity,
                "unit_price": unit_price,
                "net_value": net_value,
            }
            review_id = _create_review(conn, ticker, inferred_class.value, payment_date, candidates, "Provento B3 nao resolvido com seguranca.", payload)
            result["review_count"] += 1
            result["review_details"].append(f"Proventos Recebidos linha {row_idx}: {ticker or product} enviado para revisao")
            status = B3IncomeEventStatus.REVIEW.value
        elif _is_amortization(event_type):
            try:
                existing = _existing_income(conn, import_id, row_idx)
                if existing and existing["ledger_event_id"]:
                    ledger_event_id = existing["ledger_event_id"]
                    duplicate_reason = "import_duplicate"
                else:
                    ledger_event_id, duplicate_reason = _auto_create_amortization(conn, portfolio_id, asset_id, payment_date, net_value)
                if ledger_event_id and duplicate_reason != "import_duplicate":
                    result["auto_events_created"] += 1
                    status = B3IncomeEventStatus.LEDGER_EVENT_CREATED.value
                elif ledger_event_id:
                    status = existing["status"] if existing else B3IncomeEventStatus.LEDGER_EVENT_CREATED.value
                if duplicate_reason == "duplicate":
                    result["duplicates"] += 1
                    result["duplicate_details"].append(
                        f"Proventos Recebidos linha {row_idx}: amortizacao ja existia no ledger para {ticker or product} em {payment_date}"
                    )
            except Exception as exc:
                status = B3IncomeEventStatus.LEDGER_ERROR.value
                result["errors"].append(f"Proventos Recebidos linha {row_idx}: amortizacao nao lancada no ledger: {exc}")
        income_values.update({"status": status, "ledger_event_id": ledger_event_id, "review_id": review_id})
        inserted = _upsert_income(conn, income_values)
        if inserted:
            result["imported_incomes"] += 1
        else:
            result["duplicates"] += 1
            result["duplicate_details"].append(
                f"Proventos Recebidos linha {row_idx}: {ticker or product} ja processado neste arquivo"
            )
    return result


def _merge_counts(target: dict, source: dict) -> None:
    for key in ("total_rows", "imported_prices", "imported_incomes", "auto_events_created", "duplicates", "review_count"):
        target[key] = target.get(key, 0) + source.get(key, 0)
    target.setdefault("duplicate_details", []).extend(source.get("duplicate_details", []))
    target.setdefault("errors", []).extend(source.get("errors", []))
    target.setdefault("review_details", []).extend(source.get("review_details", []))


def import_b3_monthly_file(conn: sqlite3.Connection, portfolio_id: int, source: SourceFile) -> dict:
    reference_month, reference_date = _month_from_filename(source.filename)
    import_id = _upsert_import(conn, portfolio_id, source.filename, reference_month, reference_date)
    # Some B3 workbooks report a stale read-only dimension (A1:A1). Loading
    # normally lets openpyxl discover the real used range.
    workbook = openpyxl.load_workbook(BytesIO(source.content), read_only=False, data_only=True)
    result = {
        "filename": source.filename,
        "reference_month": reference_month,
        "reference_date": reference_date,
        "total_rows": 0,
        "imported_prices": 0,
        "imported_incomes": 0,
        "auto_events_created": 0,
        "duplicates": 0,
        "duplicate_details": [],
        "review_count": 0,
        "review_details": [],
        "errors": [],
    }

    sheet_specs = [
        ("Posição - Ações", AssetClass.ACAO, "PRECO DE FECHAMENTO", True, "CODIGO DE NEGOCIACAO", "CNPJ DA EMPRESA", None, False),
        ("Posição - Fundos", AssetClass.FII, "PRECO DE FECHAMENTO", True, "CODIGO DE NEGOCIACAO", "CNPJ DO FUNDO", None, False),
        ("Posição - Renda Fixa", AssetClass.DEBENTURE, "VALOR ATUALIZADO MTM", False, "CODIGO", None, "VENCIMENTO", True),
        ("Posição - Tesouro Direto", AssetClass.TESOURO_DIRETO, "VALOR ATUALIZADO", False, None, None, "VENCIMENTO", False),
    ]
    fixed_income_positions: list[dict] = []
    for sheet_name, asset_class, value_header, is_unit, ticker_header, cnpj_header, maturity_header, fixed_only in sheet_specs:
        ws = _worksheet_by_title(workbook, sheet_name)
        if ws is None:
            result["errors"].append(f"Aba ausente: {sheet_name}")
            continue
        partial = _process_market_sheet(
            conn,
            ws,
            import_id=import_id,
            reference_month=reference_month,
            reference_date=reference_date,
            sheet_name=sheet_name,
            asset_class=asset_class,
            value_header=value_header,
            is_unit_price=is_unit,
            ticker_header=ticker_header,
            cnpj_header=cnpj_header,
            maturity_header=maturity_header,
            fixed_income_only=fixed_only,
            fixed_income_positions=fixed_income_positions if fixed_only else None,
        )
        _merge_counts(result, partial)

    ws = _worksheet_by_title(workbook, "Proventos Recebidos")
    if ws is None:
        result["errors"].append("Aba ausente: Proventos Recebidos")
    else:
        _merge_counts(
            result,
            _process_income_sheet(
                conn,
                ws,
                import_id=import_id,
                portfolio_id=portfolio_id,
                reference_month=reference_month,
                fixed_income_positions=fixed_income_positions,
            ),
        )

    conn.execute(
        """
        UPDATE b3_monthly_imports
        SET total_rows = ?,
            imported_prices = ?,
            imported_incomes = ?,
            auto_events_created = ?,
            duplicates = ?,
            review_count = ?,
            errors = ?,
            status = ?,
            updated_at = datetime('now')
        WHERE id = ?
        """,
        (
            result["total_rows"],
            result["imported_prices"],
            result["imported_incomes"],
            result["auto_events_created"],
            result["duplicates"],
            result["review_count"],
            _json_dumps(result["errors"]) if result["errors"] else None,
            normalize_b3_monthly_import_status(B3MonthlyImportStatus.PROCESSED_WITH_ERRORS.value if result["errors"] else B3MonthlyImportStatus.PROCESSED.value),
            import_id,
        ),
    )
    return result


def import_b3_monthly_batch(conn: sqlite3.Connection, portfolio_id: int, files: list[SourceFile]) -> dict:
    ordered = sorted(files, key=lambda item: _month_from_filename(item.filename)[0])
    response = {
        "portfolio_id": portfolio_id,
        "files_processed": 0,
        "total_rows": 0,
        "imported_prices": 0,
        "imported_incomes": 0,
        "auto_events_created": 0,
        "duplicates": 0,
        "review_count": 0,
        "errors": [],
        "files": [],
    }
    for source in ordered:
        try:
            file_result = import_b3_monthly_file(conn, portfolio_id, source)
        except Exception as exc:
            reference_month, reference_date = _month_from_filename(source.filename)
            file_result = {
                "filename": source.filename,
                "reference_month": reference_month,
                "reference_date": reference_date,
                "total_rows": 0,
                "imported_prices": 0,
                "imported_incomes": 0,
                "auto_events_created": 0,
                "duplicates": 0,
                "duplicate_details": [],
                "review_count": 0,
                "review_details": [],
                "errors": [str(exc)],
            }
        response["files_processed"] += 1
        response["files"].append(file_result)
        for key in ("total_rows", "imported_prices", "imported_incomes", "auto_events_created", "duplicates", "review_count"):
            response[key] += file_result[key]
        response["errors"].extend(f"{file_result['filename']}: {err}" for err in file_result.get("errors", []))
    return response


def sanitize_b3_monthly_import(
    conn: sqlite3.Connection,
    portfolio_id: int,
    reference_month: str,
    remove_manual_resolutions: bool = False,
) -> dict:
    if not re.fullmatch(r"\d{4}-\d{2}", reference_month or ""):
        raise ValueError("Mes de referencia deve seguir o formato YYYY-MM.")
    _year, month = (int(part) for part in reference_month.split("-"))
    if month < 1 or month > 12:
        raise ValueError("Mes de referencia deve seguir o formato YYYY-MM.")

    import_rows = conn.execute(
        """
        SELECT id FROM b3_monthly_imports
        WHERE portfolio_id = ? AND reference_month = ?
        ORDER BY id
        """,
        (portfolio_id, reference_month),
    ).fetchall()
    import_ids = [row["id"] for row in import_rows]
    if not import_ids:
        return {
            "portfolio_id": portfolio_id,
            "reference_month": reference_month,
            "imports_removed": 0,
            "market_prices_removed": 0,
            "income_events_removed": 0,
            "ledger_events_cancelled": 0,
            "manual_resolutions_removed": 0,
        }

    placeholders = ",".join("?" for _ in import_ids)
    market_prices_removed = conn.execute(
        f"SELECT COUNT(*) AS count FROM b3_market_prices WHERE import_id IN ({placeholders})",
        import_ids,
    ).fetchone()["count"]
    income_events_removed = conn.execute(
        f"SELECT COUNT(*) AS count FROM b3_income_events WHERE import_id IN ({placeholders})",
        import_ids,
    ).fetchone()["count"]
    ledger_rows = conn.execute(
        f"""
        SELECT DISTINCT ledger_event_id
        FROM b3_income_events
        WHERE import_id IN ({placeholders})
          AND ledger_event_id IS NOT NULL
        ORDER BY ledger_event_id
        """,
        import_ids,
    ).fetchall()
    ledger_event_ids = [row["ledger_event_id"] for row in ledger_rows]
    review_rows = conn.execute(
        f"""
        SELECT DISTINCT review_id
        FROM (
            SELECT review_id FROM b3_income_events WHERE import_id IN ({placeholders}) AND review_id IS NOT NULL
            UNION
            SELECT review_id FROM b3_market_prices WHERE import_id IN ({placeholders}) AND review_id IS NOT NULL
        )
        ORDER BY review_id
        """,
        [*import_ids, *import_ids],
    ).fetchall()

    cancelled = []
    if ledger_event_ids:
        cancelled = event_service.delete_events_bulk(conn, ledger_event_ids)
    for row in review_rows:
        asset_service.resolve_match_review(conn, row["review_id"])

    manual_resolutions_removed = 0
    if remove_manual_resolutions:
        cur = conn.execute(
            """
            DELETE FROM b3_income_resolution_decisions
            WHERE portfolio_id = ? AND reference_month = ?
            """,
            (portfolio_id, reference_month),
        )
        manual_resolutions_removed = cur.rowcount if cur.rowcount is not None else 0

    conn.execute(
        f"DELETE FROM b3_monthly_imports WHERE id IN ({placeholders})",
        import_ids,
    )

    return {
        "portfolio_id": portfolio_id,
        "reference_month": reference_month,
        "imports_removed": len(import_ids),
        "market_prices_removed": market_prices_removed,
        "income_events_removed": income_events_removed,
        "ledger_events_cancelled": len(cancelled),
        "manual_resolutions_removed": manual_resolutions_removed,
    }
