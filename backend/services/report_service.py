"""
Report service for tax-oriented portfolio views.

Reports are derived from the event ledger by replaying historical cutoffs.
No report values are persisted here.
"""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import re
import sqlite3
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.domain.engine import (
    EngineValidationError,
    EventRecord,
    PositionState,
    process_event,
    replay_events,
)
from backend.domain.enums import AssetClass, EventType

REPORT_ASSET_CLASSES = {
    AssetClass.ACAO.value,
    AssetClass.BDR.value,
    AssetClass.CRA.value,
    AssetClass.CRI.value,
    AssetClass.CRIPTOMOEDA.value,
    AssetClass.DEBENTURE.value,
    AssetClass.ETF.value,
    AssetClass.FII.value,
    AssetClass.FI_INFRA.value,
    AssetClass.TESOURO_DIRETO.value,
    AssetClass.STOCK.value,
    AssetClass.REIT.value,
}

_ZERO = Decimal("0")
FISCAL_ASSETS_AND_RIGHTS_SECTION = "assets_and_rights"
FISCAL_TAX_EXEMPT_INCOME_SECTION = "tax_exempt_income"
FISCAL_EXCLUSIVE_TAXATION_INCOME_SECTION = "exclusive_taxation_income"
FISCAL_CAPITAL_GAINS_SECTION = "capital_gains"
DEFAULT_FISCAL_SECTIONS = [
    FISCAL_ASSETS_AND_RIGHTS_SECTION,
    FISCAL_TAX_EXEMPT_INCOME_SECTION,
    FISCAL_EXCLUSIVE_TAXATION_INCOME_SECTION,
    FISCAL_CAPITAL_GAINS_SECTION,
]
CAPITAL_GAIN_SHEETS = [
    ("B3_COMMON_15", "GC Operações Comuns"),
    ("B3_FII_FIAGRO_20", "GC FII Fiagro"),
    ("FI_INFRA_EXEMPT", "GC FI-Infra"),
    ("CRYPTO_GCAP", "GC Cripto"),
]
CAPITAL_GAIN_XLSX_COLUMNS = [
    ("month", "Mês"),
    ("realized_result", "Operações Comuns"),
    ("initial_loss_carryforward", "Resultado negativo até o mês anterior"),
    ("final_loss_carryforward", "Prejuízo a compensar"),
    ("effective_irrf", "IR fonte mês"),
    ("initial_irrf_carryforward", "IR fonte meses anteriores"),
    ("used_irrf", "IR fonte a compensar"),
    ("darf_before_minimum", "Imposto devido"),
    ("darf_estimated", "Imposto pago"),
]
INCOME_REPORT_ASSET_CLASSES = {
    AssetClass.ACAO.value,
    AssetClass.FII.value,
    AssetClass.FI_INFRA.value,
}
INCOME_TABLES = {
    "tax_exempt": {
        "title": "Rendimentos Isentos e Não Tributáveis",
        "xlsx_title": "Rendimentos Isentos",
    },
    "exclusive_taxation": {
        "title": "Rendimentos Sujeitos à Tributação Exclusiva / Definitiva",
        "xlsx_title": "Tributação Exclusiva",
    },
}
REIMBURSEMENT_PAYER_NAME = "REEMBOLSOS DE EMPRÉSTIMO DE ATIVOS"
CAPITAL_GAIN_EXEMPT_INCOME_TYPE = "Ganho líquido"
STOCK_EXEMPT_GAIN_PAYER_NAME = "Ganhos líquidos em operações no mercado à vista de ações"
FI_INFRA_EXEMPT_GAIN_PAYER_NAME = "Ganhos líquidos em operações no mercado à vista: FI-INFRA"

# Keep this table centralized so date-scoped tax rule changes can be added
# without changing the aggregation code.
INCOME_TAX_RULES = [
    {"income_type": "Dividendo", "aliases": {"Dividendo", "Dividendos"}, "table_key": "tax_exempt", "source": "b3"},
    {"income_type": "Reembolso", "aliases": {"Reembolso"}, "table_key": "tax_exempt", "source": "b3", "consolidated": True},
    {"income_type": "Redução de Capital", "aliases": {"Redução de Capital", "Reducao de Capital"}, "table_key": "tax_exempt", "source": "b3"},
    {"income_type": "Rendimento", "aliases": {"Rendimento"}, "table_key": "tax_exempt", "source": "b3"},
    {"income_type": EventType.BONIFICACAO.value, "aliases": {EventType.BONIFICACAO.value}, "table_key": "tax_exempt", "source": "ledger"},
    {"income_type": EventType.AMORTIZACAO.value, "aliases": {EventType.AMORTIZACAO.value}, "table_key": "tax_exempt", "source": "ledger_b3"},
    {"income_type": "Juros Sobre Capital Próprio", "aliases": {"Juros Sobre Capital Próprio", "Juros sobre Capital Próprio"}, "table_key": "exclusive_taxation", "source": "b3"},
    {"income_type": "Correção Selic", "aliases": {"Correção Selic", "Correcao Selic"}, "table_key": "exclusive_taxation", "source": "b3"},
]


def _norm_label(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    return normalized.encode("ascii", "ignore").decode("ascii").strip().lower()


INCOME_RULE_BY_ALIAS = {
    _norm_label(alias): rule
    for rule in INCOME_TAX_RULES
    for alias in rule["aliases"]
}


def _money(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _quantity(value: Decimal) -> str:
    normalized = value.normalize()
    _, _, exponent = normalized.as_tuple()
    if exponent >= 0 or abs(exponent) < 2:
        return str(normalized.quantize(Decimal("0.01")))
    return str(normalized)


def _xlsx_decimal(value: str) -> Decimal:
    return Decimal(value)


def fiscal_report_filename(year: int, portfolio_name: str) -> str:
    normalized = unicodedata.normalize("NFKD", portfolio_name or "")
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii").lower()
    slug = re.sub(r"[^a-z0-9]+", "_", ascii_name)
    slug = re.sub(r"_+", "_", slug).strip("_") or "carteira"
    return f"relatorio-fiscal-{year}-{slug}.xlsx"


def _event_brl_value(row: sqlite3.Row) -> Decimal:
    return Decimal(row["event_value_brl"] if row["event_value_brl"] is not None else row["event_value"])


def _income_row_id(table_key: str, income_type: str, asset_id: int | str, ticker: str | None) -> str:
    safe_ticker = re.sub(r"[^A-Za-z0-9]+", "_", ticker or "").strip("_").lower()
    return f"{table_key}:{_norm_label(income_type).replace(' ', '_')}:{asset_id}:{safe_ticker}"


def _empty_income_tables() -> dict[str, dict]:
    return {
        key: {
            "key": key,
            "title": config["title"],
            "rows": [],
            "total": _money(_ZERO),
        }
        for key, config in INCOME_TABLES.items()
    }


def _add_income_amount(
    grouped: dict[tuple[str, str, str, int | str, str, str], Decimal],
    *,
    table_key: str,
    income_type: str,
    asset_id: int | str,
    ticker: str | None,
    payer_cnpj: str | None,
    payer_name: str | None,
    amount: Decimal,
) -> None:
    if amount == _ZERO:
        return
    if income_type == "Reembolso":
        asset_id = "reimbursement"
        ticker = ""
        payer_cnpj = ""
        payer_name = REIMBURSEMENT_PAYER_NAME
    key = (table_key, income_type, ticker or "", asset_id, payer_cnpj or "", payer_name or "")
    grouped[key] += amount


def _add_capital_gain_exempt_income(
    conn: sqlite3.Connection,
    portfolio_id: int,
    year: int,
    grouped: dict[tuple[str, str, str, int | str, str, str], Decimal],
) -> None:
    from backend.services import capital_gain_report_service

    report = capital_gain_report_service.list_capital_gains(conn, portfolio_id, year, include_neutral_months=True)
    stock_exempt_gain = _ZERO
    fi_infra_exempt_gain = _ZERO

    for month in report["months"]:
        for regime in month["regimes"]:
            if regime["regime"] == "B3_COMMON_15":
                stock_exempt_gain += sum(
                    (
                        Decimal(asset["exempt_gain"])
                        for asset in regime["assets"]
                        if asset["asset_class"] == AssetClass.ACAO.value
                    ),
                    _ZERO,
                )
            elif regime["regime"] == "FI_INFRA_EXEMPT":
                fi_infra_exempt_gain += Decimal(regime["exempt_gain"])

    if stock_exempt_gain > _ZERO:
        _add_income_amount(
            grouped,
            table_key="tax_exempt",
            income_type=CAPITAL_GAIN_EXEMPT_INCOME_TYPE,
            asset_id="stock_exempt_capital_gain",
            ticker=None,
            payer_cnpj=None,
            payer_name=STOCK_EXEMPT_GAIN_PAYER_NAME,
            amount=stock_exempt_gain,
        )

    if fi_infra_exempt_gain > _ZERO:
        _add_income_amount(
            grouped,
            table_key="tax_exempt",
            income_type=CAPITAL_GAIN_EXEMPT_INCOME_TYPE,
            asset_id="fi_infra_exempt_capital_gain",
            ticker=None,
            payer_cnpj=None,
            payer_name=FI_INFRA_EXEMPT_GAIN_PAYER_NAME,
            amount=fi_infra_exempt_gain,
        )


def list_income_report(conn: sqlite3.Connection, portfolio_id: int, year: int) -> dict:
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    class_params = sorted(INCOME_REPORT_ASSET_CLASSES)
    grouped: dict[tuple[str, str, str, int | str, str, str], Decimal] = defaultdict(Decimal)

    ledger_rules = [
        rule for rule in INCOME_TAX_RULES
        if rule["source"] in {"ledger", "ledger_b3"}
    ]
    ledger_types = sorted({rule["income_type"] for rule in ledger_rules})
    ledger_rows = conn.execute(
        f"""
        SELECT
            e.id,
            e.asset_id,
            e.event_type,
            e.event_date,
            e.event_value,
            e.event_value_brl,
            e.duplicate_flag,
            a.asset_class,
            a.name,
            a.cnpj,
            (
                SELECT ticker
                FROM asset_tickers
                WHERE asset_id = e.asset_id
                  AND valid_until IS NULL
                ORDER BY valid_from DESC
                LIMIT 1
            ) AS current_ticker
        FROM events e
        JOIN assets a ON a.id = e.asset_id
        WHERE e.portfolio_id = ?
          AND e.event_date BETWEEN ? AND ?
          AND e.is_cancelled = 0
          AND e.is_storno = 0
          AND a.merged_into_asset_id IS NULL
          AND a.asset_class IN ({",".join("?" for _ in class_params)})
          AND e.event_type IN ({",".join("?" for _ in ledger_types)})
        ORDER BY current_ticker, e.event_date, e.sequence_num
        """,
        (portfolio_id, start_date, end_date, *class_params, *ledger_types),
    ).fetchall()

    ledger_rows_by_income_key = defaultdict(list)
    for row in ledger_rows:
        rule = INCOME_RULE_BY_ALIAS.get(_norm_label(row["event_type"]))
        if not rule:
            continue
        income_key = (row["asset_id"], _norm_label(rule["income_type"]), row["event_date"], _money(_event_brl_value(row)))
        ledger_rows_by_income_key[income_key].append((row, rule))

    active_ledger_income_keys = set()
    for income_key, keyed_rows in ledger_rows_by_income_key.items():
        active_ledger_income_keys.add(income_key)
        row, rule = next(
            ((candidate, candidate_rule) for candidate, candidate_rule in keyed_rows if not candidate["duplicate_flag"]),
            keyed_rows[0],
        )
        _add_income_amount(
            grouped,
            table_key=rule["table_key"],
            income_type=rule["income_type"],
            asset_id=row["asset_id"],
            ticker=row["current_ticker"],
            payer_cnpj=row["cnpj"],
            payer_name=row["name"],
            amount=_event_brl_value(row),
        )

    b3_rows = conn.execute(
        f"""
        SELECT
            i.id,
            i.asset_id,
            i.payment_date,
            i.event_type,
            i.ticker,
            i.net_value,
            i.ledger_event_id,
            a.name,
            a.cnpj,
            (
                SELECT ticker
                FROM asset_tickers
                WHERE asset_id = i.asset_id
                  AND valid_until IS NULL
                ORDER BY valid_from DESC
                LIMIT 1
            ) AS current_ticker
        FROM b3_income_events i
        JOIN assets a ON a.id = i.asset_id
        WHERE i.portfolio_id = ?
          AND i.payment_date BETWEEN ? AND ?
          AND i.asset_id IS NOT NULL
          AND i.status != 'review'
          AND i.ledger_event_id IS NULL
          AND a.merged_into_asset_id IS NULL
          AND a.asset_class IN ({",".join("?" for _ in class_params)})
        ORDER BY current_ticker, i.payment_date, i.id
        """,
        (portfolio_id, start_date, end_date, *class_params),
    ).fetchall()

    for row in b3_rows:
        rule = INCOME_RULE_BY_ALIAS.get(_norm_label(row["event_type"]))
        if not rule:
            continue
        if rule["source"] == "ledger":
            continue
        if (row["asset_id"], _norm_label(rule["income_type"]), row["payment_date"], _money(Decimal(row["net_value"] or "0"))) in active_ledger_income_keys:
            continue
        _add_income_amount(
            grouped,
            table_key=rule["table_key"],
            income_type=rule["income_type"],
            asset_id=row["asset_id"],
            ticker=row["current_ticker"] or row["ticker"],
            payer_cnpj=row["cnpj"],
            payer_name=row["name"],
            amount=Decimal(row["net_value"] or "0"),
        )

    _add_capital_gain_exempt_income(conn, portfolio_id, year, grouped)

    tables = _empty_income_tables()
    for (table_key, income_type, ticker, asset_id, payer_cnpj, payer_name), value in sorted(
        grouped.items(),
        key=lambda item: (item[0][0], _norm_label(item[0][1]), item[0][2] == "", item[0][2], str(item[0][3])),
    ):
        row = {
            "id": _income_row_id(table_key, income_type, asset_id, ticker),
            "ticker": ticker or None,
            "payer_cnpj": payer_cnpj or None,
            "payer_name": payer_name or None,
            "income_type": income_type,
            "value": _money(value),
        }
        tables[table_key]["rows"].append(row)

    for table in tables.values():
        total = sum((Decimal(row["value"]) for row in table["rows"]), _ZERO)
        table["total"] = _money(total)

    return {
        "portfolio_id": portfolio_id,
        "year": year,
        "tables": [tables["tax_exempt"], tables["exclusive_taxation"]],
    }


def _replay(records: list[EventRecord]) -> PositionState:
    try:
        return replay_events(records)
    except EngineValidationError:
        state = PositionState()
        for event in records:
            process_event(event, state, skip_validation=True)
        return state


def _rows_to_records(rows: list[sqlite3.Row]) -> list[EventRecord]:
    return [
        EventRecord(
            id=row["event_id"],
            event_type=EventType(row["event_type"]),
            event_date=row["event_date"],
            quantity=Decimal(row["quantity"]),
            event_value=Decimal(row["event_value"]),
            event_value_brl=Decimal(row["event_value_brl"]) if row["event_value_brl"] is not None else None,
            sequence_num=row["sequence_num"],
            is_cancelled=bool(row["is_cancelled"]),
            is_storno=bool(row["is_storno"]),
        )
        for row in rows
    ]


def list_assets_and_rights(conn: sqlite3.Connection, portfolio_id: int, year: int) -> dict:
    previous_cutoff = f"{year - 1}-12-31"
    current_cutoff = f"{year}-12-31"
    class_params = sorted(REPORT_ASSET_CLASSES)

    rows = conn.execute(
        f"""
        SELECT
            e.id AS event_id,
            e.asset_id,
            e.event_type,
            e.event_date,
            e.quantity,
            e.event_value,
            e.event_value_brl,
            e.sequence_num,
            e.is_cancelled,
            e.is_storno,
            a.asset_class,
            a.name,
            a.cnpj,
            (
                SELECT ticker
                FROM asset_tickers
                WHERE asset_id = e.asset_id
                  AND valid_until IS NULL
                ORDER BY valid_from DESC
                LIMIT 1
            ) AS current_ticker
        FROM events e
        JOIN assets a ON a.id = e.asset_id
        WHERE e.portfolio_id = ?
          AND e.event_date <= ?
          AND a.merged_into_asset_id IS NULL
          AND a.asset_class IN ({",".join("?" for _ in class_params)})
        ORDER BY a.asset_class, current_ticker, e.asset_id, e.event_date, e.sequence_num
        """,
        (portfolio_id, current_cutoff, *class_params),
    ).fetchall()

    grouped: dict[int, list[sqlite3.Row]] = defaultdict(list)
    for row in rows:
        grouped[row["asset_id"]].append(row)

    report_rows = []
    for asset_id, asset_rows in grouped.items():
        records = _rows_to_records(asset_rows)
        previous_state = _replay([record for record in records if record.event_date <= previous_cutoff])
        current_state = _replay(records)

        if (
            previous_state.quantity == _ZERO
            and previous_state.total_cost == _ZERO
            and current_state.quantity == _ZERO
            and current_state.total_cost == _ZERO
        ):
            continue

        first = asset_rows[0]
        report_rows.append(
            {
                "asset_id": asset_id,
                "asset_class": first["asset_class"],
                "ticker": first["current_ticker"],
                "quantity": _quantity(current_state.quantity),
                "name": first["name"],
                "cnpj": first["cnpj"],
                "previous_year_cost": _money(previous_state.total_cost),
                "current_year_cost": _money(current_state.total_cost),
            }
        )

    return {
        "portfolio_id": portfolio_id,
        "year": year,
        "previous_cutoff": previous_cutoff,
        "current_cutoff": current_cutoff,
        "rows": report_rows,
    }


def _style_header(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill


def _fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 34)


def _target_sheet(workbook: Workbook):
    if len(workbook.sheetnames) == 1 and workbook.active.max_row == 1 and workbook.active.max_column == 1 and workbook.active["A1"].value is None:
        worksheet = workbook.active
        worksheet.delete_rows(1)
        return worksheet
    return workbook.create_sheet()


def _append_assets_and_rights_sheet(workbook: Workbook, conn: sqlite3.Connection, portfolio_id: int, year: int) -> None:
    report = list_assets_and_rights(conn, portfolio_id, year)
    previous_label = f"Situação em {report['previous_cutoff'][8:10]}/{report['previous_cutoff'][5:7]}/{report['previous_cutoff'][0:4]}"
    current_label = f"Situação em {report['current_cutoff'][8:10]}/{report['current_cutoff'][5:7]}/{report['current_cutoff'][0:4]}"

    worksheet = _target_sheet(workbook)
    worksheet.title = "Bens e Direitos"
    worksheet.append(["CLASSE", "TICKER", "QUANTIDADE", "NOME ATIVO", "CNPJ", previous_label, current_label])
    _style_header(worksheet)

    for row in report["rows"]:
        worksheet.append(
            [
                row["asset_class"] or "",
                row["ticker"] or "",
                _xlsx_decimal(row["quantity"]),
                row["name"] or "",
                row["cnpj"] or "",
                _xlsx_decimal(row["previous_year_cost"]),
                _xlsx_decimal(row["current_year_cost"]),
            ]
        )
        current_row = worksheet.max_row
        worksheet[f"C{current_row}"].number_format = "#,##0.00########"
        worksheet[f"F{current_row}"].number_format = "#,##0.00"
        worksheet[f"G{current_row}"].number_format = "#,##0.00"

    _fit_columns(worksheet)


def _append_income_sheet(workbook: Workbook, table: dict, title: str) -> None:
    worksheet = _target_sheet(workbook)
    worksheet.title = title
    worksheet.append(["TICKER", "CNPJ DA FONTE PAGADORA", "NOME DA FONTE PAGADORA", "TIPO", "VALOR"])
    _style_header(worksheet)

    for row in table["rows"]:
        worksheet.append(
            [
                row["ticker"] or "",
                row["payer_cnpj"] or "",
                row["payer_name"] or "",
                row["income_type"],
                _xlsx_decimal(row["value"]),
            ]
        )
        worksheet[f"E{worksheet.max_row}"].number_format = "#,##0.00"

    _fit_columns(worksheet)


def _format_year_month(value: str) -> str:
    return f"{value[5:7]}/{value[0:4]}"


def _capital_gain_paid_xlsx_value(row: dict, paid_confirmations: set[tuple[str, str]]) -> Decimal:
    manual_tax_paid = row.get("manual_tax_paid")
    if manual_tax_paid is not None and Decimal(manual_tax_paid) > _ZERO:
        return _xlsx_decimal(manual_tax_paid)
    if (row["year_month"], row["regime"]) in paid_confirmations:
        return _xlsx_decimal(row["darf_before_minimum"])
    return _ZERO


def _append_capital_gain_sheets(workbook: Workbook, conn: sqlite3.Connection, portfolio_id: int, year: int) -> None:
    from backend.services import capital_gain_report_service, tax_service

    report = capital_gain_report_service.list_capital_gains(conn, portfolio_id, year, include_january_snapshot=True)
    paid_confirmations = {
        (row["year_month"], row["regime"])
        for row in tax_service.list_capital_gain_darf_payment_confirmations(conn, portfolio_id, year)
    }
    rows_by_regime: dict[str, list[tuple[dict, dict]]] = defaultdict(list)
    for month in report["months"]:
        for row in month["regimes"]:
            row["year_month"] = month["year_month"]
            rows_by_regime[row["regime"]].append((month, row))

    for regime, title in CAPITAL_GAIN_SHEETS:
        worksheet = _target_sheet(workbook)
        worksheet.title = title
        worksheet.append([label for _, label in CAPITAL_GAIN_XLSX_COLUMNS])
        _style_header(worksheet)

        for month, row in rows_by_regime.get(regime, []):
            worksheet.append(
                [
                    _format_year_month(month["year_month"]),
                    *[
                        _capital_gain_paid_xlsx_value(row, paid_confirmations)
                        if field == "darf_estimated"
                        else _xlsx_decimal(row[field])
                        for field, _ in CAPITAL_GAIN_XLSX_COLUMNS
                        if field != "month"
                    ],
                ]
            )
            current_row = worksheet.max_row
            for column in range(2, len(CAPITAL_GAIN_XLSX_COLUMNS) + 1):
                worksheet.cell(row=current_row, column=column).number_format = "#,##0.00"

        _fit_columns(worksheet)


def build_fiscal_report_xlsx(
    conn: sqlite3.Connection,
    portfolio_id: int,
    year: int,
    sections: list[str] | None = None,
) -> bytes:
    requested_sections = sections or DEFAULT_FISCAL_SECTIONS
    workbook = Workbook()
    income_report = None

    for section in requested_sections:
        if section == FISCAL_ASSETS_AND_RIGHTS_SECTION:
            _append_assets_and_rights_sheet(workbook, conn, portfolio_id, year)
            continue
        if section in {FISCAL_TAX_EXEMPT_INCOME_SECTION, FISCAL_EXCLUSIVE_TAXATION_INCOME_SECTION}:
            if income_report is None:
                income_report = list_income_report(conn, portfolio_id, year)
            tables = {table["key"]: table for table in income_report["tables"]}
            table_key = "tax_exempt" if section == FISCAL_TAX_EXEMPT_INCOME_SECTION else "exclusive_taxation"
            _append_income_sheet(workbook, tables[table_key], INCOME_TABLES[table_key]["xlsx_title"])
            continue
        if section == FISCAL_CAPITAL_GAINS_SECTION:
            _append_capital_gain_sheets(workbook, conn, portfolio_id, year)
            continue
        raise ValueError(f"Unknown fiscal section: {section}")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_assets_and_rights_xlsx(conn: sqlite3.Connection, portfolio_id: int, year: int) -> bytes:
    return build_fiscal_report_xlsx(conn, portfolio_id, year, sections=[FISCAL_ASSETS_AND_RIGHTS_SECTION])
